#!/usr/bin/env python3
"""
consolidar_analise.py — Consolida triagens JSON em planilha priorizada.

Gera uma planilha única ordenada por prioridade (score composto),
com colunas de filtro por executor, meta, risco, etc.

Prioridade: Prescrição > Congestionamento > Réu preso > Outras
Score: impacto_meta × facilidade_ato
"""

import json, csv
from pathlib import Path

# Importar scoring se disponível, senão usar fallback inline
try:
    from common.scoring import (
        calcular_prioridade, nivel_prioridade,
        classificar_executor, calcular_facilidade,
    )
    HAS_SCORING = True
except ImportError:
    HAS_SCORING = False


def _fallback_prioridade(d):
    """Calcula prioridade sem o módulo scoring (para rodar standalone)."""
    risco = d.get("risco_prescricao", "SEM RISCO")
    dias = int(d.get("dias_parado", 0)) if str(d.get("dias_parado", "0")).replace("-","").isdigit() else 0
    preso = d.get("reu_preso", False)
    urg = d.get("urgencia_crime", d.get("urgencia", "MEDIA"))

    peso_presc = {"PRESCRITO": 15000, "IMINENTE": 12000, "ATENCAO": 10000, "BAIXO": 500, "SEM RISCO": 0}
    score = peso_presc.get(risco, 0)

    if dias >= 1825: score += 5000
    elif dias >= 1095: score += 4000
    elif dias >= 730: score += 3000
    elif dias >= 365: score += 2000
    elif dias >= 180: score += 1000
    else: score += 500

    if preso: score += 3000
    peso_crime = {"CRITICA": 800, "ALTA": 400, "MEDIA": 200, "BAIXA": 100}
    score += peso_crime.get(urg, 200)

    fac = int(d.get("facilidade_ato", 3))
    mult = 1.0 + (fac - 1) * 0.25
    score = int(score * mult)

    if risco in ("PRESCRITO", "IMINENTE", "ATENCAO"): meta = "Prescrição"
    elif dias >= 365: meta = "Congestionamento"
    elif preso: meta = "Réu preso"
    else: meta = "Outras"

    return score, meta


def carregar_triagens(resultados_dir: Path) -> list:
    """Carrega todos os JSONs de triagem e retorna lista unificada."""
    jsons = sorted(resultados_dir.glob("triagem_*.json"))
    if not jsons:
        print("  Nenhum arquivo de triagem encontrado.")
        return []

    todos = []
    numeros_vistos = set()

    for jp in jsons:
        try:
            dados = json.loads(jp.read_text(encoding="utf-8"))
            if isinstance(dados, list):
                for d in dados:
                    num = d.get("numero", "")
                    if num and num not in numeros_vistos:
                        numeros_vistos.add(num)
                        todos.append(d)
            elif isinstance(dados, dict):
                num = dados.get("numero", "")
                if num and num not in numeros_vistos:
                    numeros_vistos.add(num)
                    todos.append(dados)
        except Exception as e:
            print(f"  Erro lendo {jp.name}: {e}")

    return todos


def enriquecer_e_ordenar(dados: list) -> list:
    """Adiciona campos de priorização e ordena por score decrescente."""
    for d in dados:
        proximo_ato = d.get("proximo_ato", "")
        risco = d.get("risco_prescricao", "SEM RISCO")
        dias = int(d.get("dias_parado", 0)) if str(d.get("dias_parado", "0")).replace("-","").isdigit() else 0
        preso = d.get("reu_preso", False)
        if isinstance(preso, str):
            preso = preso.lower() in ("true", "sim", "s", "1")
        urg = d.get("urgencia_crime", d.get("urgencia", "MEDIA"))

        if HAS_SCORING:
            pri = calcular_prioridade(risco, dias, proximo_ato, preso, urg)
            d["score_prioridade"] = pri["score_prioridade"]
            d["meta_principal"] = pri["meta_principal"]
            if "executor" not in d or not d["executor"]:
                d["executor"] = pri["executor"]
            if "facilidade_ato" not in d or not d["facilidade_ato"]:
                d["facilidade_ato"] = pri["facilidade_ato"]
            d["nivel_prioridade"] = nivel_prioridade(pri["score_prioridade"])
        else:
            score, meta = _fallback_prioridade(d)
            d["score_prioridade"] = score
            d["meta_principal"] = meta
            if "executor" not in d: d["executor"] = "Verificar"
            if "facilidade_ato" not in d: d["facilidade_ato"] = 3

            if score >= 15000: d["nivel_prioridade"] = "URGENTÍSSIMA"
            elif score >= 10000: d["nivel_prioridade"] = "URGENTE"
            elif score >= 5000: d["nivel_prioridade"] = "ALTA"
            elif score >= 2000: d["nivel_prioridade"] = "MÉDIA"
            else: d["nivel_prioridade"] = "NORMAL"

    dados.sort(key=lambda x: -x.get("score_prioridade", 0))
    return dados


def gerar_xlsx(dados: list, path: Path):
    """Gera planilha .xlsx com formatação profissional e filtros."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  openpyxl não encontrado. Gerando CSV como fallback.")
        _fallback_csv(dados, path)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Priorização"

    # ── Estilos ──
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Arial", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Cores por nível de prioridade
    cores_nivel = {
        "URGENTÍSSIMA": (PatternFill("solid", fgColor="8B0000"), Font(name="Arial", size=10, bold=True, color="FFFFFF")),
        "URGENTE":      (PatternFill("solid", fgColor="FF0000"), Font(name="Arial", size=10, bold=True, color="FFFFFF")),
        "ALTA":         (PatternFill("solid", fgColor="FF6600"), Font(name="Arial", size=10, bold=True, color="FFFFFF")),
        "MÉDIA":        (PatternFill("solid", fgColor="FFCC00"), Font(name="Arial", size=10, bold=True)),
        "NORMAL":       (PatternFill("solid", fgColor="D9EAD3"), Font(name="Arial", size=10)),
    }

    cores_prescricao = {
        "PRESCRITO": (PatternFill("solid", fgColor="8B0000"), Font(name="Arial", size=10, bold=True, color="FFFFFF")),
        "IMINENTE":  (PatternFill("solid", fgColor="FF0000"), Font(name="Arial", size=10, bold=True, color="FFFFFF")),
        "ATENCAO":   (PatternFill("solid", fgColor="FF6600"), Font(name="Arial", size=10, bold=True, color="FFFFFF")),
        "BAIXO":     (PatternFill("solid", fgColor="92D050"), Font(name="Arial", size=10)),
        "SEM RISCO": (PatternFill("solid", fgColor="D9EAD3"), Font(name="Arial", size=10)),
    }

    cores_executor = {
        "Cartório":    PatternFill("solid", fgColor="DAEEF3"),
        "Assessoria":  PatternFill("solid", fgColor="E2EFDA"),
        "Externo":     PatternFill("solid", fgColor="FFF2CC"),
        "Juiz":        PatternFill("solid", fgColor="F2DCDB"),
        "Verificar":   PatternFill("solid", fgColor="F2F2F2"),
    }

    # ── Colunas ──
    colunas = [
        ("Nº",             "numero",            8),
        ("Prioridade",     "nivel_prioridade",  14),
        ("Score",          "score_prioridade",   9),
        ("Meta",           "meta_principal",     16),
        ("Executor",       "executor",           13),
        ("Próximo Ato",    "proximo_ato",        40),
        ("Facilidade",     "facilidade_ato",      11),
        ("Classe",         "classe",             10),
        ("Assunto",        "assunto",            25),
        ("Dias Parado",    "dias_parado",         11),
        ("Prescrição",     "risco_prescricao",   13),
        ("Réu Preso",      "reu_preso",           10),
        ("Fase",           "fase_processual",    30),
        ("Resumo",         "resumo",             55),
        ("Fundamentação",  "fundamentacao_legal", 35),
        ("Peças-Chave",    "pecas_chave",        35),
    ]

    # ── Cabeçalho ──
    for col_idx, (titulo, _, largura) in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}{len(dados) + 1}"

    # ── Dados ──
    for row_idx, d in enumerate(dados, 2):
        for col_idx, (_, campo, _) in enumerate(colunas, 1):
            valor = d.get(campo, "")

            # Converter booleano
            if campo == "reu_preso":
                if isinstance(valor, bool):
                    valor = "SIM" if valor else "NÃO"
                elif isinstance(valor, str):
                    valor = "SIM" if valor.lower() in ("true", "sim", "s", "1") else "NÃO"

            # Converter facilidade para texto legível
            if campo == "facilidade_ato":
                fac_map = {5: "5-Trivial", 4: "4-Simples", 3: "3-Médio", 2: "2-Complexo", 1: "1-Pesado"}
                valor = fac_map.get(int(valor) if str(valor).isdigit() else 3, str(valor))

            # Número do processo: só os 7 primeiros dígitos + ano para ficar legível
            if campo == "numero" and isinstance(valor, str) and len(valor) > 20:
                pass  # manter completo para filtro

            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

        # ── Cores condicionais ──
        nivel = d.get("nivel_prioridade", "NORMAL")
        if nivel in cores_nivel:
            fill, font = cores_nivel[nivel]
            ws.cell(row=row_idx, column=2).fill = fill
            ws.cell(row=row_idx, column=2).font = font

        presc = d.get("risco_prescricao", "SEM RISCO")
        if presc in cores_prescricao:
            fill, font = cores_prescricao[presc]
            ws.cell(row=row_idx, column=11).fill = fill
            ws.cell(row=row_idx, column=11).font = font

        executor = d.get("executor", "Verificar")
        if executor in cores_executor:
            ws.cell(row=row_idx, column=5).fill = cores_executor[executor]

        # Réu preso em vermelho
        preso_val = ws.cell(row=row_idx, column=12).value
        if preso_val == "SIM":
            ws.cell(row=row_idx, column=12).fill = PatternFill("solid", fgColor="FF0000")
            ws.cell(row=row_idx, column=12).font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

    # ── Linha de totais no topo (linha 2 seria dados, então usar rodapé) ──
    # Não adicionar, pois atrapalha os filtros

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  ✅ Planilha: {path} ({len(dados)} processos)")


def _fallback_csv(dados: list, path: Path):
    """Fallback CSV quando openpyxl não está disponível."""
    saida = path.with_suffix(".csv")
    campos = [
        "numero", "nivel_prioridade", "score_prioridade", "meta_principal",
        "executor", "proximo_ato", "facilidade_ato", "classe", "assunto",
        "dias_parado", "risco_prescricao", "reu_preso", "fase_processual",
        "resumo", "fundamentacao_legal", "pecas_chave",
    ]
    with open(saida, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=campos, extrasaction="ignore")
        w.writeheader()
        w.writerows(dados)
    print(f"  CSV fallback: {saida}")


def gerar_resumo(dados: list):
    """Imprime resumo estatístico da triagem."""
    total = len(dados)
    if not total:
        return

    # Por nível
    niveis = {}
    for d in dados:
        n = d.get("nivel_prioridade", "?")
        niveis[n] = niveis.get(n, 0) + 1

    # Por executor
    executores = {}
    for d in dados:
        e = d.get("executor", "?")
        executores[e] = executores.get(e, 0) + 1

    # Por meta
    metas = {}
    for d in dados:
        m = d.get("meta_principal", "?")
        metas[m] = metas.get(m, 0) + 1

    # Prescrição
    prescritos = sum(1 for d in dados if d.get("risco_prescricao") == "PRESCRITO")
    iminentes = sum(1 for d in dados if d.get("risco_prescricao") == "IMINENTE")
    atencao = sum(1 for d in dados if d.get("risco_prescricao") == "ATENCAO")

    # Réu preso
    presos = sum(1 for d in dados if d.get("reu_preso") in (True, "true", "SIM", "sim", "True"))

    print(f"\n  {'─'*50}")
    print(f"  RESUMO: {total} processos analisados")
    print(f"  {'─'*50}")

    print(f"\n  Por prioridade:")
    for n in ["URGENTÍSSIMA", "URGENTE", "ALTA", "MÉDIA", "NORMAL"]:
        if n in niveis:
            print(f"    {n}: {niveis[n]}")

    print(f"\n  Por executor (quem faz o próximo ato):")
    for e, c in sorted(executores.items(), key=lambda x: -x[1]):
        print(f"    {e}: {c}")

    print(f"\n  Por meta impactada:")
    for m, c in sorted(metas.items(), key=lambda x: -x[1]):
        print(f"    {m}: {c}")

    if prescritos or iminentes or atencao:
        print(f"\n  ⚠️  PRESCRIÇÃO:")
        if prescritos: print(f"    🔴 PRESCRITOS: {prescritos}")
        if iminentes:  print(f"    🟠 IMINENTES: {iminentes}")
        if atencao:    print(f"    🟡 ATENÇÃO: {atencao}")

    if presos:
        print(f"\n  🔒 Réus presos: {presos}")

    print()


class ConsolidarAnalise:
    """Consolida resultados de análise em planilha priorizada."""

    def __init__(self, service_dir: Path, result_dir: Path):
        self.resultados_dir = service_dir / "resultados"
        self.analises_dir = self.resultados_dir / "analises"
        self.result_dir = result_dir
        self.result_dir.mkdir(parents=True, exist_ok=True)

    def consolidar(self):
        print(f"\n{'='*60}")
        print(f"  CONSOLIDAÇÃO — Triagem Priorizada")
        print(f"{'='*60}")

        dados = carregar_triagens(self.resultados_dir)
        if not dados:
            print("  Nenhum dado para consolidar.")
            return

        print(f"  {len(dados)} processos carregados")

        dados = enriquecer_e_ordenar(dados)

        saida = self.result_dir / "triagem_processos.xlsx"
        gerar_xlsx(dados, saida)

        self._contar_analises()
        gerar_resumo(dados)

    def _contar_analises(self):
        if self.analises_dir.exists():
            mds = list(self.analises_dir.glob("*.md"))
            print(f"  📁 Análises detalhadas: {len(mds)} arquivos em resultados/analises/")
        else:
            print("  📁 Pasta analises/ não encontrada.")


if __name__ == "__main__":
    import sys
    service = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("services/analisar_processo")
    result = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("result/analisar_processo")
    ConsolidarAnalise(service, result).consolidar()
