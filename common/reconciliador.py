"""
common/reconciliador.py — Casa nomes da lista do papel (xlsx) com nomes do PJe.

Estratégia em 3 camadas, confiança decrescente:
  1. Âncora pelo número do processo (CNJ) — mais confiável
  2. Match exato após normalização (sem acentos, lowercase, tokens ordenados)
  3. Fuzzy matching com rapidfuzz para erros ortográficos

Uso:
    from common.reconciliador import Reconciliador

    rec = Reconciliador()
    rec.carregar_lista_papel("files/lista_cadastro_scc.xlsx")
    rec.carregar_pje("files/scc_info.csv")

    rec.reconciliar()
    rec.exportar_relatorio("result/reconciliacao.xlsx")

Dependências:
    pip install rapidfuzz openpyxl unidecode
"""

import re
import csv
import json
from pathlib import Path
from dataclasses import dataclass, asdict
from typing import Optional

try:
    from rapidfuzz import fuzz, process
    from unidecode import unidecode
except ImportError:
    raise ImportError("Instale: pip install rapidfuzz unidecode openpyxl")


# ── Normalização ────────────────────────────────────────────────

# Sufixos parentéticos comuns na lista do papel: "(BALCÃO)", "(APF)", "A.P"
_RE_SUFIXO = re.compile(r"\s*[\(\[][^)\]]*[\)\]]\s*|\s+A\.?P\.?\s*$|\s+APF\s*$", re.I)
_RE_CNJ = re.compile(r"\d{7}-\d{2}\.\d{4}\.\d{1}\.\d{2}\.\d{4}")
_RE_ESPACOS = re.compile(r"\s+")


def normalizar_nome(nome: str) -> str:
    """
    Normaliza nome para comparação:
      - remove sufixos parentéticos
      - remove acentos
      - lowercase
      - ordena tokens (resolve "Silva, João" vs "João Silva")
      - colapsa espaços
    """
    if not nome:
        return ""
    s = _RE_SUFIXO.sub(" ", nome)
    s = unidecode(s)
    s = _RE_ESPACOS.sub(" ", s.lower().strip())
    tokens = sorted(s.split())
    return " ".join(tokens)


def extrair_cnj(texto: str) -> Optional[str]:
    """Extrai número CNJ do texto, se houver."""
    m = _RE_CNJ.search(texto or "")
    return m.group(0) if m else None


# ── Modelo ──────────────────────────────────────────────────────

@dataclass
class Match:
    """Resultado do casamento entre um item do papel e um do PJe."""
    papel_idx: int
    papel_nome: str
    papel_processo: str
    papel_livro: str

    pje_nome: Optional[str]
    pje_processo: Optional[str]
    pje_id_processo: Optional[str]

    score: int                # 0-100
    metodo: str               # "cnj" | "exato" | "fuzzy" | "sem_match"
    revisar: bool


# ── Reconciliador ───────────────────────────────────────────────

class Reconciliador:

    THRESHOLD_AUTO = 90
    THRESHOLD_REVISAR = 70

    def __init__(self):
        self.papel: list[dict] = []
        self.pje: list[dict] = []
        self._matches: list[Match] = []

    def carregar_lista_papel(self, xlsx_path: str | Path):
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if not row or not row[2]:
                continue
            num, processo, nome, livro, etiquetado = (row + (None,) * 5)[:5]
            self.papel.append({
                "idx": i,
                "num_papel": num,
                "processo_raw": str(processo or ""),
                "processo_cnj": extrair_cnj(str(processo or "")),
                "nome": str(nome).strip(),
                "nome_norm": normalizar_nome(str(nome)),
                "livro": str(livro or "").strip(),
                "etiquetado": str(etiquetado or "").strip(),
            })
        wb.close()
        print(f"  Papel: {len(self.papel)} cadastros")

    def carregar_pje(self, csv_path: str | Path):
        with open(csv_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f, delimiter=";")
            for row in reader:
                nome = (row.get("poloPassivo") or "").strip()
                if not nome:
                    continue
                self.pje.append({
                    "id_processo": row.get("idProcesso", "").strip(),
                    "numero_processo": row.get("numeroProcesso", "").strip(),
                    "classe": row.get("classeJudicial", "").strip(),
                    "polo_passivo": nome,
                    "nome_norm": normalizar_nome(nome),
                    "assunto": row.get("assuntoPrincipal", "").strip(),
                })
        # Deduplica por (numero_processo + nome)
        vistos = set()
        unicos = []
        for r in self.pje:
            chave = (r["numero_processo"], r["nome_norm"])
            if chave not in vistos:
                vistos.add(chave)
                unicos.append(r)
        self.pje = unicos
        print(f"  PJe: {len(self.pje)} réus únicos")

    def reconciliar(self) -> list[Match]:
        matches: list[Match] = []
        pje_por_cnj: dict[str, list[dict]] = {}
        for r in self.pje:
            pje_por_cnj.setdefault(r["numero_processo"], []).append(r)
        pje_por_nome_norm = {r["nome_norm"]: r for r in self.pje}
        pje_nomes_norm = [r["nome_norm"] for r in self.pje]

        for p in self.papel:
            match = self._match_um(p, pje_por_cnj, pje_por_nome_norm, pje_nomes_norm)
            matches.append(match)

        self._matches = matches
        self._imprimir_resumo(matches)
        return matches

    def _match_um(self, papel, pje_por_cnj, pje_por_nome_norm, pje_nomes_norm) -> Match:
        base = dict(
            papel_idx=papel["idx"],
            papel_nome=papel["nome"],
            papel_processo=papel["processo_raw"],
            papel_livro=papel["livro"],
        )

        # Camada 1: CNJ
        if papel["processo_cnj"] and papel["processo_cnj"] in pje_por_cnj:
            candidatos = pje_por_cnj[papel["processo_cnj"]]
            melhor = max(
                candidatos,
                key=lambda c: fuzz.token_sort_ratio(papel["nome_norm"], c["nome_norm"]),
            )
            score_nome = fuzz.token_sort_ratio(papel["nome_norm"], melhor["nome_norm"])
            return Match(
                **base,
                pje_nome=melhor["polo_passivo"],
                pje_processo=melhor["numero_processo"],
                pje_id_processo=melhor["id_processo"],
                score=max(95, int(score_nome)),
                metodo="cnj",
                revisar=score_nome < 70,
            )

        # Camada 2: nome exato
        if papel["nome_norm"] in pje_por_nome_norm:
            r = pje_por_nome_norm[papel["nome_norm"]]
            return Match(
                **base,
                pje_nome=r["polo_passivo"],
                pje_processo=r["numero_processo"],
                pje_id_processo=r["id_processo"],
                score=92,
                metodo="exato",
                revisar=False,
            )

        # Camada 3: fuzzy
        if pje_nomes_norm:
            res = process.extractOne(
                papel["nome_norm"], pje_nomes_norm, scorer=fuzz.token_sort_ratio,
            )
            if res:
                _, score, idx = res
                if score >= self.THRESHOLD_REVISAR:
                    r = self.pje[idx]
                    return Match(
                        **base,
                        pje_nome=r["polo_passivo"],
                        pje_processo=r["numero_processo"],
                        pje_id_processo=r["id_processo"],
                        score=int(score),
                        metodo="fuzzy",
                        revisar=score < self.THRESHOLD_AUTO,
                    )

        return Match(
            **base, pje_nome=None, pje_processo=None, pje_id_processo=None,
            score=0, metodo="sem_match", revisar=True,
        )

    def _imprimir_resumo(self, matches: list[Match]):
        total = len(matches)
        por_metodo: dict[str, int] = {}
        revisar = 0
        for m in matches:
            por_metodo[m.metodo] = por_metodo.get(m.metodo, 0) + 1
            if m.revisar:
                revisar += 1

        print(f"\n  ── Reconciliação ({total} cadastros do papel) ──")
        for metodo in ("cnj", "exato", "fuzzy", "sem_match"):
            n = por_metodo.get(metodo, 0)
            pct = n / total * 100 if total else 0
            print(f"    {metodo:>10}: {n:3} ({pct:5.1f}%)")
        print(f"    {'a revisar':>10}: {revisar:3}")

    def exportar_relatorio(self, xlsx_path: str | Path):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Reconciliação"

        cabecalho = [
            "Linha Papel", "Nome (Papel)", "Processo (Papel)", "Livro/Fls",
            "Nome (PJe)", "Processo (PJe)", "ID Processo PJe",
            "Score", "Método", "Revisar?",
        ]
        ws.append(cabecalho)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F3864")

        verde = PatternFill("solid", fgColor="D9EAD3")
        amarelo = PatternFill("solid", fgColor="FFF2CC")
        vermelho = PatternFill("solid", fgColor="F4CCCC")

        for i, m in enumerate(self._matches, 2):
            ws.append([
                m.papel_idx, m.papel_nome, m.papel_processo, m.papel_livro,
                m.pje_nome or "—", m.pje_processo or "—", m.pje_id_processo or "—",
                m.score, m.metodo, "SIM" if m.revisar else "—",
            ])
            fill = vermelho if m.metodo == "sem_match" else (amarelo if m.revisar else verde)
            for col in range(1, len(cabecalho) + 1):
                ws.cell(row=i, column=col).fill = fill

        for col_letter, width in zip("ABCDEFGHIJ", [12, 35, 30, 18, 35, 30, 14, 8, 12, 10]):
            ws.column_dimensions[col_letter].width = width
        ws.freeze_panes = "A2"

        Path(xlsx_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(xlsx_path)
        print(f"  Relatório: {xlsx_path}")

    def exportar_json(self, json_path: str | Path):
        """Exporta dict numero_processo → match para uso em downstream."""
        data = {
            m.papel_processo: {
                "papel_nome": m.papel_nome,
                "papel_livro": m.papel_livro,
                "pje_nome": m.pje_nome,
                "score": m.score,
                "metodo": m.metodo,
                "revisar": m.revisar,
            }
            for m in self._matches if m.papel_processo
        }
        Path(json_path).parent.mkdir(parents=True, exist_ok=True)
        Path(json_path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    import sys
    rec = Reconciliador()
    rec.carregar_lista_papel(sys.argv[1] if len(sys.argv) > 1 else "files/lista_cadastro_scc.xlsx")
    rec.carregar_pje(sys.argv[2] if len(sys.argv) > 2 else "files/scc_info.csv")
    rec.reconciliar()
    rec.exportar_relatorio(sys.argv[3] if len(sys.argv) > 3 else "result/reconciliacao_papel_pje.xlsx")
