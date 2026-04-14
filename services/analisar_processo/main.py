import sys, re, json, csv
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from common.fila_base import FilaBase
from common.checkpoint import CheckpointManager
from common.sessao import SessaoManager
from common.utils import DIR_RESULT

SERVICE_DIR = Path(__file__).parent
RESULT_DIR = DIR_RESULT / "analisar_processo"
ANALISES_DIR = SERVICE_DIR / "resultados" / "analises"
ANALISES_DIR.mkdir(parents=True, exist_ok=True)

ROTA_BASE = {
    "APOrd": [
        "knowledge/processual/ritos_processuais.md (secao APOrd)",
        "knowledge/criminal/prescricao.md",
        "knowledge/modelos/despachos.md",
    ],
    "IP": [
        "knowledge/processual/ritos_processuais.md (secao IP)",
        "knowledge/criminal/prescricao.md",
        "knowledge/modelos/despachos.md",
    ],
    "TCO": [
        "knowledge/processual/ritos_processuais.md (secao Sumarissima)",
        "knowledge/criminal/prescricao.md",
        "knowledge/modelos/despachos.md",
    ],
    "Juri": [
        "knowledge/processual/ritos_processuais.md (secao Juri)",
        "knowledge/criminal/prescricao.md",
        "knowledge/criminal/crimes_pessoa.md (secao Homicidio)",
        "knowledge/modelos/decisoes.md",
        "knowledge/modelos/despachos.md",
    ],
    "APSum": [
        "knowledge/processual/ritos_processuais.md (secao APSum)",
        "knowledge/criminal/prescricao.md",
        "knowledge/modelos/despachos.md",
    ],
    "APSumss": [
        "knowledge/processual/ritos_processuais.md (secao Sumarissima)",
        "knowledge/criminal/prescricao.md",
        "knowledge/modelos/despachos.md",
    ],
}

ROTA_DEFAULT = [
    "knowledge/processual/ritos_processuais.md",
    "knowledge/criminal/prescricao.md",
    "knowledge/modelos/despachos.md",
    "knowledge/processual/competencia.md",
]

KEYWORDS_LEIS = [
    (r"tráfico|drogas|entorpecente|cocaína|maconha|crack", "knowledge/leis/drogas.md"),
    (r"violência doméstica|maria da penha|mulher|protetiva|VD", "knowledge/leis/maria_penha.md"),
    (r"arma|porte ilegal|posse.*arma|munição|desarmamento", "knowledge/leis/armas.md"),
    (r"trânsito|embriaguez|CTB|volante|atropelamento", "knowledge/leis/transito.md"),
    (r"peculato|corrupção|concussão|prevaricação|funcionário", "knowledge/leis/adm_publica.md"),
    (r"organização criminosa|facção|associação criminosa", "knowledge/criminal/organizacao_criminosa.md"),
    (r"homicídio|feminicídio|lesão corporal|estupro|cárcere", "knowledge/criminal/crimes_pessoa.md"),
    (r"roubo|furto|latrocínio|extorsão|estelionato|receptação", "knowledge/criminal/crimes_patrimonio.md"),
]


def _detectar_extras(assunto):
    extras = []
    for pattern, arquivo in KEYWORDS_LEIS:
        if re.search(pattern, assunto, re.IGNORECASE):
            extras.append(arquivo)
    return extras


def _montar_rota(classe, assunto):
    base = list(ROTA_BASE.get(classe, ROTA_DEFAULT))
    extras = _detectar_extras(assunto)
    vistos = set()
    rota = []
    for arq in base + extras:
        nome = arq.split(" (")[0]
        if nome not in vistos:
            vistos.add(nome)
            rota.append(arq)
    return rota


def _fmt_rota(rota):
    return "\n".join(f"  {i}. {arq}" for i, arq in enumerate(rota, 1))


def _num_to_filename(numero):
    return numero.replace(".", "_").replace("-", "_")


INSTRUCAO_SAIDA = """
## FORMATO DE SAIDA (OBRIGATORIO)

Para CADA processo, gere DOIS arquivos:

### 1. Ficha de triagem (JSON) — uma linha por processo
Adicione ao arquivo services/analisar_processo/resultados/triagem_{cmd}.json:
```json
[
  {{
    "numero": "0000000-00.2020.8.05.0216",
    "classe": "APOrd",
    "assunto": "Roubo",
    "dias_parado": 450,
    "urgencia": "CRITICA",
    "fase_processual": "Alegacoes apresentadas",
    "proximo_ato": "Minutar sentenca",
    "risco_prescricao": "ATENCAO",
    "resumo": "Reu denunciado por roubo majorado. AIJ realizada. Alegacoes de ambas as partes apresentadas. Concluso para sentenca."
  }}
]
```

### 2. Analise completa (Markdown) — um arquivo por processo
Salve em services/analisar_processo/resultados/analises/{numero_arquivo}.md:

```markdown
# Analise — {numero_processo}

## Dados
- **Classe**: ...
- **Assunto**: ...
- **Dias parado**: ...
- **Urgencia**: CRITICA / ALTA / MEDIA / BAIXA
- **Risco de prescricao**: PRESCRITO / IMINENTE / ATENCAO / BAIXO / SEM RISCO

## Situacao Atual
[Resumo detalhado da situacao do processo, citando pecas e paginas]

## Fase Processual
[Fase exata em que o processo se encontra]

## Diagnostico
[Analise juridica detalhada do que precisa ser feito e por que]

## Proximo Ato
[Ato concreto e especifico — nunca "dar andamento"]

## Modelo de Despacho
```
[Texto completo do despacho pronto para o juiz assinar]
```

## Fundamentacao Legal
[Artigos e dispositivos aplicaveis]

## Pecas-Chave
[Lista das pecas relevantes com paginas: "Denuncia (p.3-5, Num. XXXX)", etc.]

## Observacoes
[Qualquer alerta adicional: nulidade, excesso de prazo, custodia, etc.]
```

IMPORTANTE:
- O nome do arquivo markdown DEVE ser o numero do processo com _ no lugar de . e -
  Exemplo: 0000770-14.2020.8.05.0216 -> 0000770_14_2020_8_05_0216.md
- Crie a pasta analises/ se nao existir
- O JSON de triagem deve ter um objeto por processo no array
"""


class FilaAnalise(FilaBase):
    SERVICE_NAME = "analisar_processo"
    BATCH_COM_PDF = 3
    BATCH_SEM_PDF = 15
    CLASSE_PARA_PROMPT = {
        "APOrd": "prompt_APOrd.md", "IP": "prompt_IP.md", "TCO": "prompt_TCO.md",
        "Juri": "prompt_Juri.md", "APSum": "prompt_APSum.md", "APSumss": "prompt_APSumss.md",
    }

    def gerar_comando_com_pdf(self, n, procs, prompt):
        arqs = " ".join(f"textos_extraidos/{p['txt_arquivo']}" for p in procs)
        nums = " ".join(p["numero"] for p in procs)
        resumo = "\n".join(
            f"  {p['numero']} | {p['assunto']} | {p['dias_parado']}d | {p['urgencia']}"
            for p in procs
        )
        assuntos = " ".join(p.get("assunto", "") for p in procs)
        rota = _fmt_rota(_montar_rota(procs[0]["classe"], assuntos))
        instrucao = INSTRUCAO_SAIDA.replace("{cmd}", f"{n:03d}")

        return f"""# === CMD {n:03d} [{procs[0]['classe']}] [{len(procs)} procs] ===
# Ao concluir: python run.py analise marcar {n} {nums}

Leia services/analisar_processo/prompts/{prompt}

Rota de leitura:
{rota}
  +reu preso -> knowledge/processual/cautelares_e_prisao.md
  +recurso pendente -> knowledge/processual/recursos.md
  +minutar sentenca -> knowledge/criminal/dosimetria.md + knowledge/modelos/sentencas.md

Processos:
{resumo}

Arquivos: {arqs}
{instrucao}"""

    def gerar_comando_sem_pdf(self, n, procs, prompt):
        tab = "| Numero | Classe | Assunto | Dias | Ultima Mov. |\n|---|---|---|---|---|\n"
        tab += "\n".join(
            f"| {p['numero']} | {p['classe']} | {p['assunto']} | {p['dias_parado']} | {p['ultima_mov'][:40]} |"
            for p in procs
        )
        nums = " ".join(p["numero"] for p in procs)
        assuntos = " ".join(p.get("assunto", "") for p in procs)
        rota = _fmt_rota(_montar_rota(procs[0]["classe"], assuntos))
        instrucao = INSTRUCAO_SAIDA.replace("{cmd}", f"{n:03d}")

        return f"""# === CMD {n:03d} [{procs[0]['classe']}] [{len(procs)} SEM PDF] ===
# Ao concluir: python run.py analise marcar {n} {nums}

Leia services/analisar_processo/prompts/{prompt}

Rota de leitura:
{rota}

Analise LIMITADA (sem PDF):

{tab}
{instrucao}"""


class ConsolidarAnalise:
    def __init__(self, service_dir, result_dir):
        self.resultados_dir = service_dir / "resultados"
        self.analises_dir = self.resultados_dir / "analises"
        self.result_dir = result_dir
        self.result_dir.mkdir(parents=True, exist_ok=True)

    def consolidar(self):
        print(f"{'='*60}\n  CONSOLIDACAO\n{'='*60}")
        self._gerar_xlsx()
        self._contar_analises()

    def _gerar_xlsx(self):
        jsons = sorted(self.resultados_dir.glob("triagem_*.json"))
        if not jsons:
            print("  Nenhum arquivo de triagem encontrado.")
            return

        todos = []
        for jp in jsons:
            try:
                dados = json.loads(jp.read_text(encoding="utf-8"))
                if isinstance(dados, list):
                    todos.extend(dados)
                else:
                    todos.append(dados)
            except Exception as e:
                print(f"  Erro lendo {jp.name}: {e}")

        if not todos:
            print("  Nenhum registro de triagem.")
            return

        ord_urg = {"CRITICA": 0, "ALTA": 1, "MEDIA": 2, "BAIXA": 3}
        ord_presc = {"PRESCRITO": 0, "IMINENTE": 1, "ATENCAO": 2, "BAIXO": 3, "SEM RISCO": 4}
        todos.sort(key=lambda x: (
            ord_urg.get(x.get("urgencia", "BAIXA"), 4),
            ord_presc.get(x.get("risco_prescricao", "SEM RISCO"), 5),
            -int(x.get("dias_parado", 0)) if str(x.get("dias_parado", "0")).isdigit() else 0,
        ))

        saida = self.result_dir / "triagem_processos.xlsx"
        self._xlsx(todos, saida)
        print(f"  Planilha: {saida} ({len(todos)} processos)")

    def _xlsx(self, dados, path):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            print("  pip install openpyxl")
            self._fallback_csv(dados, path)
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Triagem"

        hf = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        hfill = PatternFill("solid", fgColor="2F5496")
        ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
        nf = Font(name="Arial", size=10)
        bd = Border(*[Side(style="thin", color="D9D9D9")] * 4)

        cores_urg = {
            "CRITICA": PatternFill("solid", fgColor="FF0000"),
            "ALTA": PatternFill("solid", fgColor="FF6600"),
            "MEDIA": PatternFill("solid", fgColor="FFCC00"),
            "BAIXA": PatternFill("solid", fgColor="92D050"),
        }
        cores_presc = {
            "PRESCRITO": PatternFill("solid", fgColor="FF0000"),
            "IMINENTE": PatternFill("solid", fgColor="FF6600"),
            "ATENCAO": PatternFill("solid", fgColor="FFCC00"),
            "BAIXO": PatternFill("solid", fgColor="92D050"),
            "SEM RISCO": PatternFill("solid", fgColor="D9EAD3"),
        }
        font_branca = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        font_preta = Font(name="Arial", size=10, bold=True)

        cols = [
            ("A", "Numero", 30),
            ("B", "Classe", 12),
            ("C", "Assunto", 25),
            ("D", "Dias", 8),
            ("E", "Urgencia", 12),
            ("F", "Prescricao", 14),
            ("G", "Fase", 25),
            ("H", "Proximo Ato", 35),
            ("I", "Resumo", 50),
        ]

        for c, t, w in cols:
            cell = ws[f"{c}1"]
            cell.value = t
            cell.font = hf
            cell.fill = hfill
            cell.alignment = ha
            cell.border = bd
            ws.column_dimensions[c].width = w

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:I{len(dados)+1}"

        for i, d in enumerate(dados, 2):
            urg = d.get("urgencia", "BAIXA")
            presc = d.get("risco_prescricao", "SEM RISCO")

            row = [
                d.get("numero", ""),
                d.get("classe", ""),
                d.get("assunto", ""),
                d.get("dias_parado", ""),
                urg,
                presc,
                d.get("fase_processual", ""),
                d.get("proximo_ato", ""),
                d.get("resumo", ""),
            ]

            for j, v in enumerate(row):
                cell = ws.cell(row=i, column=j + 1, value=v)
                cell.font = nf
                cell.border = bd
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            cell_urg = ws.cell(row=i, column=5)
            if urg in cores_urg:
                cell_urg.fill = cores_urg[urg]
                cell_urg.font = font_branca if urg in ("CRITICA", "ALTA") else font_preta

            cell_presc = ws.cell(row=i, column=6)
            if presc in cores_presc:
                cell_presc.fill = cores_presc[presc]
                cell_presc.font = font_branca if presc in ("PRESCRITO", "IMINENTE") else font_preta

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)

    def _fallback_csv(self, dados, path):
        saida = path.with_suffix(".csv")
        campos = ["numero", "classe", "assunto", "dias_parado", "urgencia",
                   "risco_prescricao", "fase_processual", "proximo_ato", "resumo"]
        with open(saida, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=campos, extrasaction="ignore")
            w.writeheader()
            w.writerows(dados)
        print(f"  CSV fallback: {saida}")

    def _contar_analises(self):
        if self.analises_dir.exists():
            mds = list(self.analises_dir.glob("*.md"))
            print(f"  Analises individuais: {len(mds)} arquivos em resultados/analises/")
        else:
            print("  Pasta analises/ nao encontrada.")


def executar(comando, args=None):
    args = args or []
    if comando == "fila":
        FilaAnalise(SERVICE_DIR).gerar()
    elif comando == "status":
        FilaAnalise(SERVICE_DIR).status()
    elif comando == "analisar":
        SessaoManager(SERVICE_DIR / "checkpoint.json").inicio()
        print(f"  Cole comandos de: {SERVICE_DIR / 'comandos_claude_code.txt'}")
    elif comando == "pausa":
        SessaoManager(SERVICE_DIR / "checkpoint.json").fim(SERVICE_DIR / "fila.json")
    elif comando == "marcar":
        if not args:
            print("  USO: marcar <NUM> [PROCESSOS...]")
            return
        CheckpointManager(SERVICE_DIR / "checkpoint.json").marcar_concluido(
            int(args[0]), args[1:], f"resultados/triagem_{int(args[0]):03d}.json"
        )
    elif comando == "consolidar":
        ConsolidarAnalise(SERVICE_DIR, RESULT_DIR).consolidar()
    elif comando == "reset":
        CheckpointManager(SERVICE_DIR / "checkpoint.json").reset()
        for f in [SERVICE_DIR / "fila.json", SERVICE_DIR / "comandos_claude_code.txt"]:
            if f.exists():
                f.unlink()
        print("  Reset OK.")
    else:
        print(f"  Desconhecido: {comando}")
