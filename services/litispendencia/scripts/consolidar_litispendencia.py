#!/usr/bin/env python3
"""
services/litispendencia/scripts/consolidar_litispendencia.py

Lê todos os grupo_*.json de resultados/grupos/ e produz planilha xlsx
com 4 abas: Resumo, Pares Litispendência, Falsos Positivos, Estatísticas.

Uso:
    python -m services.litispendencia.scripts.consolidar_litispendencia
    python run.py litispendencia consolidar
"""

import json
import sys
from collections import Counter
from pathlib import Path

RAIZ = Path(__file__).resolve().parents[3]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from common.utils import DIR_RESULT

SERVICE_DIR = RAIZ / "services" / "litispendencia"
GRUPOS_DIR = SERVICE_DIR / "resultados" / "grupos"
RESULT_DIR = DIR_RESULT / "litispendencia"
SAIDA_PADRAO = RESULT_DIR / "triagem_litispendencia.xlsx"


# Cores para coloração condicional
CORES_CLASSIF = {
    "LITISPENDENCIA_TOTAL":   "FF6B6B",  # vermelho claro
    "LITISPENDENCIA_PARCIAL": "FFB870",  # laranja
    "COISA_JULGADA":          "C0392B",  # vermelho escuro
    "CONEXAO":                "F1C40F",  # amarelo
    "CONTINENCIA":            "F39C12",  # laranja escuro
    "CAUSAS_DISTINTAS":       "82C99B",  # verde
    "INDEFINIDO":             "BDC3C7",  # cinza
}
CORES_PRIORIDADE = {
    "URGENTE": "C0392B",
    "ALTA":    "E67E22",
    "MEDIA":   "F1C40F",
    "BAIXA":   "95A5A6",
}
CORES_CONFIANCA = {
    "ALTA":  "27AE60",
    "MEDIA": "F39C12",
    "BAIXA": "C0392B",
}

ORDEM_PRIORIDADE = {"URGENTE": 0, "ALTA": 1, "MEDIA": 2, "BAIXA": 3, "": 4}


def carregar_analises() -> list[dict]:
    if not GRUPOS_DIR.exists():
        return []
    analises = []
    for json_path in sorted(GRUPOS_DIR.glob("grupo_*.json")):
        try:
            analises.append(json.loads(json_path.read_text(encoding="utf-8")))
        except json.JSONDecodeError as e:
            print(f"  ⚠️  {json_path.name}: JSON inválido ({e})")
    return analises


def gerar_aba_resumo(wb, analises):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ws = wb.active
    ws.title = "Resumo por Grupo"

    cabecalho = [
        "Group ID", "Aba Origem", "N Processos",
        "Classificação", "Confiança", "Prioridade",
        "Executor", "Facilidade",
        "Pares Identif.", "Falsos Pos.", "Coisa Julg.",
        "Processo Mais Antigo", "Providência Sugerida",
        "Observações",
    ]

    cab_font = Font(bold=True, color="FFFFFF", size=10)
    cab_fill = PatternFill("solid", fgColor="1F3864")
    cab_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borda = Border(*[Side(style="thin", color="D9D9D9")] * 4)

    for i, titulo in enumerate(cabecalho, 1):
        c = ws.cell(row=1, column=i, value=titulo)
        c.font = cab_font
        c.fill = cab_fill
        c.alignment = cab_align
        c.border = borda

    ws.row_dimensions[1].height = 30
    larguras = [12, 22, 8, 22, 10, 11, 14, 11, 12, 11, 11, 24, 50, 50]
    for i, l in enumerate(larguras, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = l
    ws.freeze_panes = "A2"

    # Ordenar: prioridade asc, depois grupos maiores primeiro
    analises_ord = sorted(
        analises,
        key=lambda a: (
            ORDEM_PRIORIDADE.get(a.get("prioridade", ""), 4),
            -a.get("n_processos", 0),
        ),
    )

    for ri, a in enumerate(analises_ord, 2):
        classif = a.get("classificacao_final", "")
        confianca = a.get("confianca", "")
        prioridade = a.get("prioridade", "")

        valores = [
            a.get("group_id", ""),
            a.get("aba_origem", ""),
            a.get("n_processos", 0),
            classif,
            confianca,
            prioridade,
            a.get("executor", ""),
            a.get("facilidade_ato", ""),
            len(a.get("pares_litispendencia", [])),
            len(a.get("processos_distintos", [])),
            len(a.get("processos_coisa_julgada", [])),
            a.get("processo_mais_antigo", ""),
            a.get("providencia_sugerida", ""),
            a.get("observacoes", ""),
        ]
        for ci, v in enumerate(valores, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = borda

        # Coloração condicional
        if classif in CORES_CLASSIF:
            ws.cell(row=ri, column=4).fill = PatternFill("solid", fgColor=CORES_CLASSIF[classif])
            ws.cell(row=ri, column=4).font = Font(bold=True, color="FFFFFF" if classif in ("COISA_JULGADA",) else "000000")
        if confianca in CORES_CONFIANCA:
            ws.cell(row=ri, column=5).fill = PatternFill("solid", fgColor=CORES_CONFIANCA[confianca])
            ws.cell(row=ri, column=5).font = Font(bold=True, color="FFFFFF")
        if prioridade in CORES_PRIORIDADE:
            ws.cell(row=ri, column=6).fill = PatternFill("solid", fgColor=CORES_PRIORIDADE[prioridade])
            ws.cell(row=ri, column=6).font = Font(bold=True, color="FFFFFF")

    if analises_ord:
        from openpyxl.utils import get_column_letter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cabecalho))}{len(analises_ord) + 1}"


def gerar_aba_pares(wb, analises):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ws = wb.create_sheet("Pares Litispendência")
    cabecalho = [
        "Group ID", "Tipo", "Processo A", "Processo B", "Outros",
        "Justificativa",
    ]
    cab_font = Font(bold=True, color="FFFFFF", size=10)
    cab_fill = PatternFill("solid", fgColor="1F3864")
    borda = Border(*[Side(style="thin", color="D9D9D9")] * 4)

    for i, t in enumerate(cabecalho, 1):
        c = ws.cell(row=1, column=i, value=t)
        c.font = cab_font
        c.fill = cab_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda

    larguras = [12, 22, 26, 26, 30, 70]
    for i, l in enumerate(larguras, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = l
    ws.freeze_panes = "A2"

    ri = 2
    for a in analises:
        for par in a.get("pares_litispendencia", []):
            procs = par.get("processos", [])
            valores = [
                a.get("group_id", ""),
                par.get("tipo", ""),
                procs[0] if len(procs) > 0 else "",
                procs[1] if len(procs) > 1 else "",
                ", ".join(procs[2:]) if len(procs) > 2 else "",
                par.get("justificativa", ""),
            ]
            for ci, v in enumerate(valores, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = borda

            tipo = par.get("tipo", "")
            if tipo in CORES_CLASSIF:
                ws.cell(row=ri, column=2).fill = PatternFill("solid", fgColor=CORES_CLASSIF[tipo])
                ws.cell(row=ri, column=2).font = Font(bold=True)
            ri += 1

    if ri > 2:
        from openpyxl.utils import get_column_letter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cabecalho))}{ri - 1}"


def gerar_aba_falsos_positivos(wb, analises):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ws = wb.create_sheet("Falsos Positivos")
    cabecalho = ["Group ID", "Aba Origem", "Processo", "Justificativa"]
    cab_font = Font(bold=True, color="FFFFFF", size=10)
    cab_fill = PatternFill("solid", fgColor="1F3864")
    borda = Border(*[Side(style="thin", color="D9D9D9")] * 4)

    for i, t in enumerate(cabecalho, 1):
        c = ws.cell(row=1, column=i, value=t)
        c.font = cab_font
        c.fill = cab_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda

    larguras = [12, 22, 26, 80]
    for i, l in enumerate(larguras, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = l
    ws.freeze_panes = "A2"

    ri = 2
    for a in analises:
        for dp in a.get("processos_distintos", []):
            valores = [
                a.get("group_id", ""),
                a.get("aba_origem", ""),
                dp.get("numero", ""),
                dp.get("justificativa", ""),
            ]
            for ci, v in enumerate(valores, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = borda
            ri += 1


def gerar_aba_estatisticas(wb, analises):
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.create_sheet("Estatísticas")

    cab_font = Font(bold=True, size=11)

    # Contagens
    por_classif = Counter(a.get("classificacao_final", "") for a in analises)
    por_prioridade = Counter(a.get("prioridade", "") for a in analises)
    por_confianca = Counter(a.get("confianca", "") for a in analises)
    por_aba = Counter(a.get("aba_origem", "") for a in analises)

    total_pares = sum(len(a.get("pares_litispendencia", [])) for a in analises)
    total_distintos = sum(len(a.get("processos_distintos", [])) for a in analises)
    total_coisa_julg = sum(len(a.get("processos_coisa_julgada", [])) for a in analises)

    linhas = [
        ("Total de grupos analisados", len(analises)),
        ("", ""),
        ("── Por classificação ──", ""),
    ]
    for k in ["LITISPENDENCIA_TOTAL", "LITISPENDENCIA_PARCIAL", "COISA_JULGADA",
              "CONEXAO", "CONTINENCIA", "CAUSAS_DISTINTAS", "INDEFINIDO"]:
        linhas.append((k, por_classif.get(k, 0)))
    linhas.append(("", ""))
    linhas.append(("── Por prioridade ──", ""))
    for k in ["URGENTE", "ALTA", "MEDIA", "BAIXA"]:
        linhas.append((k, por_prioridade.get(k, 0)))
    linhas.append(("", ""))
    linhas.append(("── Por confiança ──", ""))
    for k in ["ALTA", "MEDIA", "BAIXA"]:
        linhas.append((k, por_confianca.get(k, 0)))
    linhas.append(("", ""))
    linhas.append(("── Por aba de origem ──", ""))
    for k, v in por_aba.most_common():
        linhas.append((k, v))
    linhas.append(("", ""))
    linhas.append(("── Totais agregados ──", ""))
    linhas.append(("Pares de litispendência identificados", total_pares))
    linhas.append(("Falsos positivos (processos distintos)", total_distintos))
    linhas.append(("Processos com coisa julgada", total_coisa_julg))

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 12

    for i, (k, v) in enumerate(linhas, 1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        if k.startswith("──"):
            ws.cell(row=i, column=1).font = cab_font
            ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor="D9E1F2")


def consolidar(saida: Path | None = None):
    saida = saida or SAIDA_PADRAO

    print(f"\n  Lendo: {GRUPOS_DIR}")
    analises = carregar_analises()
    if not analises:
        print(f"  ✗ Nenhuma análise encontrada em {GRUPOS_DIR}")
        print(f"     Rode primeiro: python auto_analisar_litispendencia.py")
        return

    print(f"  {len(analises)} grupos analisados\n")

    try:
        from openpyxl import Workbook
    except ImportError:
        print("  ✗ openpyxl não instalado. Rode: pip install openpyxl")
        return

    wb = Workbook()
    gerar_aba_resumo(wb, analises)
    gerar_aba_pares(wb, analises)
    gerar_aba_falsos_positivos(wb, analises)
    gerar_aba_estatisticas(wb, analises)

    saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(saida)

    # Resumo no terminal
    por_clas = Counter(a.get("classificacao_final", "") for a in analises)
    print(f"  ── Resumo ──")
    print(f"  Total grupos:              {len(analises)}")
    print(f"  LITISPENDENCIA_TOTAL:      {por_clas.get('LITISPENDENCIA_TOTAL', 0)}")
    print(f"  LITISPENDENCIA_PARCIAL:    {por_clas.get('LITISPENDENCIA_PARCIAL', 0)}")
    print(f"  COISA_JULGADA:             {por_clas.get('COISA_JULGADA', 0)}")
    print(f"  CONEXAO:                   {por_clas.get('CONEXAO', 0)}")
    print(f"  CONTINENCIA:               {por_clas.get('CONTINENCIA', 0)}")
    print(f"  CAUSAS_DISTINTAS:          {por_clas.get('CAUSAS_DISTINTAS', 0)}")
    print(f"  INDEFINIDO:                {por_clas.get('INDEFINIDO', 0)}")
    print(f"\n  ✓ Planilha: {saida}\n")


if __name__ == "__main__":
    consolidar()
