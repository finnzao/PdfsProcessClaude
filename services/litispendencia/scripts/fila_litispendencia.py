#!/usr/bin/env python3
"""
services/litispendencia/scripts/fila_litispendencia.py

Gera fila de comandos para o Claude Code analisar grupos de
litispendência a partir da planilha xlsx do filtro pré-classificador.

Estratégia adaptativa:
  - Grupos pequenos (≤3 procs): empacotados em CMDs com até 6 procs no total
  - Grupos grandes (≥6 procs): isolados, 1 grupo por CMD
  - Grupos médios (4-5 procs): isolados também

Salva (no padrão do projeto):
  services/litispendencia/fila.json
  services/litispendencia/comandos_claude_code.txt

Pula grupos já listados em controle_grupos.json (a menos que --forcar).

Uso:
    python -m services.litispendencia.scripts.fila_litispendencia
    python -m services.litispendencia.scripts.fila_litispendencia --xlsx files/litispendencia_2.xlsx
    python -m services.litispendencia.scripts.fila_litispendencia --abas "Litispendência,Coisa Julgada"
    python -m services.litispendencia.scripts.fila_litispendencia --forcar
"""

import argparse
import json
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

RAIZ = Path(__file__).resolve().parents[3]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

from common.utils import DIR_FILES, agora_iso

SERVICE_DIR = RAIZ / "services" / "litispendencia"
FILA_PATH = SERVICE_DIR / "fila.json"
CMDS_PATH = SERVICE_DIR / "comandos_claude_code.txt"
CONTROLE_PATH = SERVICE_DIR / "controle_grupos.json"
TEXTOS_DIR = RAIZ / "textos_extraidos"

# Empacotamento adaptativo
ALVO_PROCS_POR_CMD = 6       # alvo para grupos pequenos
LIMITE_GRUPO_GRANDE = 6      # 6+ procs = sempre isolado

# Mapa de aba → prefixo do group_id (chaves normalizadas: sem acento, lower)
PREFIXO_ABA = {
    "litispendencia": "lit",
    "coisa julgada": "cj",
    "filtro estrito": "estrito",
    "filtro medio": "medio",
    "filtro amplo": "amplo",
}

ABAS_DEFAULT = ["Litispendência", "Coisa Julgada"]

# Variantes antigas (com emoji) → nome canônico atual. Mantém compat.
ALIASES_ABA = {
    "⭐ Litispendência": "Litispendência",
    "⚠ Coisa Julgada": "Coisa Julgada",
}


def _normalizar_chave(s: str) -> str:
    """Remove acentos, espaços extras e emojis, lowercase. Para casamento robusto."""
    if not s:
        return ""
    # Remove caracteres não-ASCII tipo emoji
    sem_emoji = "".join(c for c in s if c.isprintable() and ord(c) < 0x2700)
    nfkd = unicodedata.normalize("NFKD", sem_emoji)
    sem_acento = "".join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", sem_acento).strip().lower()


def _prefixo_para_aba(aba_nome: str) -> str:
    return PREFIXO_ABA.get(_normalizar_chave(aba_nome), "grp")


def carregar_controle():
    if CONTROLE_PATH.exists():
        try:
            return json.loads(CONTROLE_PATH.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            print(f"  ⚠️  {CONTROLE_PATH.name} corrompido — começando do zero")
    return {
        "atualizado_em": agora_iso(),
        "total_analisados": 0,
        "grupos": {},
    }


def num_para_md(numero_cnj):
    """0001234-56.2024.8.05.0216 → 0001234_56_2024_8_05_0216.md"""
    return numero_cnj.replace(".", "_").replace("-", "_") + ".md"


def md_existe(numero_cnj):
    return (TEXTOS_DIR / num_para_md(numero_cnj)).exists()


def _resolver_aba(nome_pedido: str, abas_workbook: list[str]) -> str | None:
    """Resolve o nome da aba lidando com alias (emoji antigo) e acentos.

    Retorna o nome exato como está no workbook, ou None se não houver match.
    """
    # 1. Match exato
    if nome_pedido in abas_workbook:
        return nome_pedido

    # 2. Alias direto (emoji antigo → novo nome)
    if nome_pedido in ALIASES_ABA:
        alvo = ALIASES_ABA[nome_pedido]
        if alvo in abas_workbook:
            return alvo

    # 3. Match normalizado (sem acento, sem emoji, lower)
    pedido_norm = _normalizar_chave(nome_pedido)
    for nome_real in abas_workbook:
        if _normalizar_chave(nome_real) == pedido_norm:
            return nome_real

    return None


def _detectar_linha_cabecalho(ws, max_busca: int = 5) -> tuple[int, list[str]]:
    """Encontra a linha com 'Grupo' e 'Nº Processo' nas primeiras N linhas.

    Retorna (numero_linha_1based, lista_cabecalho_strings).
    A linha 1 da planilha costuma ser um título descritivo mesclado; o
    cabeçalho real está na linha 2 nesta planilha. Detectamos por conteúdo
    em vez de assumir posição.
    """
    for n_linha, row in enumerate(ws.iter_rows(min_row=1, max_row=max_busca, values_only=True), 1):
        cab = [str(c).strip() if c is not None else "" for c in row]
        cab_norm = [_normalizar_chave(c) for c in cab]
        tem_grupo = any(c == "grupo" for c in cab_norm)
        tem_processo = any("processo" in c for c in cab_norm)
        if tem_grupo and tem_processo:
            return n_linha, cab
    return 1, []


def _achar_coluna(cabecalho: list[str], *nomes_alvo: str) -> int | None:
    """Procura coluna pelo nome (normalizado). Aceita match exato ou substring."""
    alvos = [_normalizar_chave(n) for n in nomes_alvo]
    cab_norm = [_normalizar_chave(c) for c in cabecalho]

    # 1. Match exato
    for i, c in enumerate(cab_norm):
        if c in alvos:
            return i
    # 2. Match por substring (qualquer alvo dentro do cabeçalho da coluna)
    for i, c in enumerate(cab_norm):
        for alvo in alvos:
            if alvo and alvo in c:
                return i
    return None


def ler_xlsx(xlsx_path: Path, abas: list[str]) -> list[dict]:
    """Lê grupos das abas pedidas. Detecta cabeçalho dinamicamente.

    Estrutura esperada da planilha:
      - Linha 1: título descritivo mesclado
      - Linha 2: cabeçalho (Grupo, Polo Ativo, Polo Passivo, Classe, Assunto,
                 Nº Processo, Data Chegada, Tarefa Atual, Status, ...)
      - Linha 3+: uma linha por processo, agrupados pela coluna 'Grupo'

    Cada item retornado: {
      'group_id': 'lit_001',
      'aba_origem': 'Litispendência',
      'processos': ['0001234-...', ...],
      'n_processos': N,
      'partes_amostra': 'AUTOR vs RÉU',
      'classe_amostra': 'CumSenFaz',
      'assunto_amostra': 'Execução Contratual',
      'status_amostra': 'ARQUIVADO; ATIVO; ...'
    }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  ✗ openpyxl não instalado. Rode: pip install openpyxl")
        sys.exit(1)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    grupos_out = []
    contadores = {}

    re_cnj = re.compile(r"\d{7}-\d{2}\.\d{4}\.\d\.\d{2}\.\d{4}")

    for aba_pedida in abas:
        aba_real = _resolver_aba(aba_pedida, wb.sheetnames)
        if not aba_real:
            print(f"  ⚠️  Aba não encontrada: {aba_pedida!r}")
            print(f"     Disponíveis: {', '.join(repr(s) for s in wb.sheetnames)}")
            continue
        if aba_real != aba_pedida:
            print(f"  ℹ️  Aba {aba_pedida!r} mapeada para {aba_real!r}")

        ws = wb[aba_real]
        prefixo = _prefixo_para_aba(aba_real)
        contadores.setdefault(prefixo, 0)

        # Detecta linha do cabeçalho e identifica colunas pelos nomes
        linha_cab, cabecalho = _detectar_linha_cabecalho(ws)
        if not cabecalho:
            print(f"  ⚠️  Cabeçalho não detectado em {aba_real!r}, pulando")
            continue

        col_grupo = _achar_coluna(cabecalho, "Grupo")
        col_proc = _achar_coluna(cabecalho, "Nº Processo", "N Processo", "Numero do Processo", "Processo")
        col_polo_ativo = _achar_coluna(cabecalho, "Polo Ativo")
        col_polo_passivo = _achar_coluna(cabecalho, "Polo Passivo")
        col_classe = _achar_coluna(cabecalho, "Classe")
        col_assunto = _achar_coluna(cabecalho, "Assunto")
        col_status = _achar_coluna(cabecalho, "Status")

        if col_grupo is None or col_proc is None:
            print(f"  ⚠️  {aba_real!r}: faltam colunas 'Grupo' ou 'Nº Processo' (cabeçalho linha {linha_cab})")
            continue

        # Agrupar pelas linhas de dados (após o cabeçalho)
        agrupados: dict[str, dict] = {}
        for row in ws.iter_rows(min_row=linha_cab + 1, values_only=True):
            if not row or all(c is None for c in row):
                continue

            grupo_val = row[col_grupo] if col_grupo < len(row) else None
            proc_val = row[col_proc] if col_proc < len(row) else None
            if not grupo_val or not proc_val:
                continue

            grupo_chave = str(grupo_val).strip()
            cnj_match = re_cnj.search(str(proc_val))
            if not cnj_match:
                continue
            cnj = cnj_match.group(0)

            g = agrupados.setdefault(grupo_chave, {
                "processos": [],
                "polos_ativos": [],
                "polos_passivos": [],
                "classes": [],
                "assuntos": [],
                "status": [],
            })
            if cnj not in g["processos"]:
                g["processos"].append(cnj)

            def _add(lista, idx):
                if idx is not None and idx < len(row) and row[idx]:
                    v = str(row[idx]).strip()
                    if v and v not in lista:
                        lista.append(v)

            _add(g["polos_ativos"], col_polo_ativo)
            _add(g["polos_passivos"], col_polo_passivo)
            _add(g["classes"], col_classe)
            _add(g["assuntos"], col_assunto)
            _add(g["status"], col_status)

        # Materializa grupos válidos (≥ 2 processos)
        # Ordena por nome (Grupo 1, Grupo 2, ...) com fallback alfabético
        def _chave_ordem(nome):
            m = re.search(r"(\d+)", nome)
            return (0, int(m.group(1))) if m else (1, nome)

        for grupo_chave in sorted(agrupados.keys(), key=_chave_ordem):
            g = agrupados[grupo_chave]
            if len(g["processos"]) < 2:
                continue

            contadores[prefixo] += 1
            group_id = f"{prefixo}_{contadores[prefixo]:03d}"

            ativos = " ↔ ".join(g["polos_ativos"][:2]) or ""
            passivos = " ↔ ".join(g["polos_passivos"][:2]) or ""
            partes_amostra = f"{ativos} vs {passivos}" if ativos or passivos else ""

            grupos_out.append({
                "group_id": group_id,
                "aba_origem": aba_real,
                "grupo_planilha": grupo_chave,
                "processos": g["processos"],
                "n_processos": len(g["processos"]),
                "partes_amostra": partes_amostra[:160],
                "classe_amostra": "; ".join(g["classes"][:3]),
                "assunto_amostra": "; ".join(g["assuntos"][:3]),
                "status_amostra": "; ".join(g["status"][:5]),
            })

    wb.close()
    return grupos_out


def empacotar_adaptativo(grupos: list[dict]) -> list[list[dict]]:
    """Empacota grupos em CMDs respeitando ALVO_PROCS_POR_CMD.

    Grupos grandes (≥ LIMITE_GRUPO_GRANDE) sempre isolados.
    Grupos pequenos empilhados até ALVO_PROCS_POR_CMD processos.
    """
    cmds = []
    buffer = []
    procs_buffer = 0

    for g in grupos:
        n = g["n_processos"]

        # Grupo grande: flush buffer + grupo isolado
        if n >= LIMITE_GRUPO_GRANDE:
            if buffer:
                cmds.append(buffer)
                buffer = []
                procs_buffer = 0
            cmds.append([g])
            continue

        # Grupo médio/pequeno: cabe no buffer?
        if procs_buffer + n > ALVO_PROCS_POR_CMD and buffer:
            cmds.append(buffer)
            buffer = []
            procs_buffer = 0

        buffer.append(g)
        procs_buffer += n

    if buffer:
        cmds.append(buffer)

    return cmds


def gerar_texto_cmd(num_cmd: int, grupos_do_cmd: list[dict]) -> str:
    """Gera o texto de um CMD com 1+ grupos."""
    group_ids = " ".join(g["group_id"] for g in grupos_do_cmd)
    n_procs_total = sum(g["n_processos"] for g in grupos_do_cmd)

    # Tabela de grupos no comando
    linhas_tabela = []
    for g in grupos_do_cmd:
        cab_linha = f"  {g['group_id']:<12} [{g['aba_origem']} / {g.get('grupo_planilha', '?')}]"
        linhas_tabela.append(cab_linha)
        for cnj in g["processos"]:
            existe = "✓" if md_existe(cnj) else "✗"
            linhas_tabela.append(f"  {'':12}   - {cnj}  (md={existe})")
        if g.get("partes_amostra"):
            linhas_tabela.append(f"  {'':12}   partes: {g['partes_amostra'][:120]}")
        if g.get("classe_amostra") or g.get("assunto_amostra"):
            ca = g.get("classe_amostra", "")
            ass = g.get("assunto_amostra", "")
            linhas_tabela.append(f"  {'':12}   classe/assunto: {ca} / {ass}")
        if g.get("status_amostra"):
            linhas_tabela.append(f"  {'':12}   status: {g['status_amostra']}")
        linhas_tabela.append("")

    tabela = "\n".join(linhas_tabela).rstrip()

    # Lista de arquivos .md a ler
    todos_cnjs = []
    for g in grupos_do_cmd:
        todos_cnjs.extend(g["processos"])
    todos_cnjs_unicos = list(dict.fromkeys(todos_cnjs))  # dedup preservando ordem

    arquivos = " ".join(
        f"textos_extraidos/{num_para_md(c)}"
        for c in todos_cnjs_unicos
        if md_existe(c)
    )

    return f"""# === CMD {num_cmd:03d} [LITISPENDENCIA] [{len(grupos_do_cmd)} grupo(s), {n_procs_total} proc(s)] ===
# Grupos: {group_ids}
# Ao concluir: python run.py litispendencia marcar {num_cmd} {group_ids}

Leia services/litispendencia/prompts/prompt_litispendencia.md e siga TODAS
as regras dele, em especial:
  - Salvamento INCREMENTAL após cada grupo (não só no fim do CMD)
  - Pular grupos já listados em controle_grupos.json
  - Distinguir rótulo (classe+assunto+partes) de causa de pedir
  - Múltiplos pares possíveis dentro do mesmo grupo grande

Grupos a analisar (NESTA ORDEM):

{tabela}

Arquivos .md disponíveis para leitura:
{arquivos}

Para CADA grupo (na ordem acima):

1. Ler services/litispendencia/controle_grupos.json — pular se group_id já presente
2. Ler todos os .md disponíveis do grupo (alguns podem faltar)
3. Comparar partes, classe, assunto E causa de pedir
4. Salvar services/litispendencia/resultados/grupos/grupo_<group_id>.json
   conforme schema do prompt
5. Atualizar controle_grupos.json com o group_id
6. SÓ ENTÃO ir ao próximo grupo

LEMBRETES CRÍTICOS:
- Litispendência exige partes + pedido + CAUSA DE PEDIR idênticos
- Dois CumSenFaz entre as mesmas partes podem executar sentenças diferentes
- Em grupo grande, pode haver múltiplos pares + processos distintos
- Se faltar .md, classifique INDEFINIDO + confianca BAIXA, não invente
"""


def gerar_fila(xlsx_path: Path, abas: list[str], forcar: bool):
    if not xlsx_path.exists():
        print(f"  ✗ Planilha não encontrada: {xlsx_path}")
        print(f"     Coloque o arquivo lá ou use --xlsx <caminho>")
        sys.exit(1)

    print(f"\n  Lendo: {xlsx_path}")
    print(f"  Abas: {', '.join(abas)}")

    grupos = ler_xlsx(xlsx_path, abas)
    if not grupos:
        print(f"  ✗ Nenhum grupo extraído das abas pedidas.")
        print(f"     Dicas:")
        print(f"       - Verifique se os nomes das abas estão corretos (--abas '...')")
        print(f"       - Verifique se há ao menos 2 processos por valor de 'Grupo'")
        sys.exit(1)

    print(f"\n  Grupos encontrados: {len(grupos)}")
    por_aba = {}
    for g in grupos:
        por_aba[g["aba_origem"]] = por_aba.get(g["aba_origem"], 0) + 1
    for aba, n in por_aba.items():
        print(f"    {aba}: {n} grupos")

    # Filtrar já analisados
    controle = carregar_controle()
    ja_analisados = set(controle.get("grupos", {}).keys())

    if forcar:
        print(f"  --forcar: ignorando controle_grupos.json ({len(ja_analisados)} grupos)")
    else:
        antes = len(grupos)
        grupos = [g for g in grupos if g["group_id"] not in ja_analisados]
        if antes != len(grupos):
            print(f"  Já analisados (pulados): {antes - len(grupos)}")

    if not grupos:
        print(f"\n  Nada a fazer. Use --forcar para reprocessar tudo.")
        return

    # Estatística de distribuição
    pequenos = sum(1 for g in grupos if g["n_processos"] <= 3)
    medios = sum(1 for g in grupos if 4 <= g["n_processos"] <= 5)
    grandes = sum(1 for g in grupos if g["n_processos"] >= 6)
    print(f"\n  Distribuição:")
    print(f"    Pequenos (2-3 procs):  {pequenos}")
    print(f"    Médios   (4-5 procs):  {medios}")
    print(f"    Grandes  (6+ procs):   {grandes}")

    # Cobertura de .md
    todos_procs = [c for g in grupos for c in g["processos"]]
    com_md = sum(1 for c in todos_procs if md_existe(c))
    print(f"\n  Cobertura de textos_extraidos/: {com_md}/{len(todos_procs)} processos têm .md")

    # Empacotar adaptativo
    cmds_pacotes = empacotar_adaptativo(grupos)

    print(f"\n  CMDs gerados: {len(cmds_pacotes)} (vs {len(grupos)} se fosse 1:1)")

    # Gerar comandos
    cmds_meta = []
    with open(CMDS_PATH, "w", encoding="utf-8") as f:
        f.write(f"# litispendencia — {len(cmds_pacotes)} comandos — "
                f"{datetime.now():%d/%m/%Y %H:%M}\n")
        f.write(f"# Total: {len(grupos)} grupos, "
                f"{sum(g['n_processos'] for g in grupos)} processos\n\n")

        for i, pacote in enumerate(cmds_pacotes, 1):
            texto = gerar_texto_cmd(i, pacote)
            f.write(texto + "\n\n")
            cmds_meta.append({
                "num": i,
                "grupos": [g["group_id"] for g in pacote],
                "processos": [c for g in pacote for c in g["processos"]],
                "n_grupos": len(pacote),
                "n_procs": sum(g["n_processos"] for g in pacote),
                "tipo": "LITISPENDENCIA",
            })

    # fila.json no formato compatível com FilaBase/auto_analisar
    fila = {
        "gerado_em": agora_iso(),
        "service": "litispendencia",
        "abas": abas,
        "total_grupos": len(grupos),
        "total_processos": sum(g["n_processos"] for g in grupos),
        "total_comandos": len(cmds_pacotes),
        "grupos_detalhe": [
            {
                "group_id": g["group_id"],
                "aba_origem": g["aba_origem"],
                "grupo_planilha": g.get("grupo_planilha", ""),
                "n_processos": g["n_processos"],
                "processos": g["processos"],
            }
            for g in grupos
        ],
        "comandos": cmds_meta,
    }
    FILA_PATH.write_text(
        json.dumps(fila, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(f"\n  ✓ Fila salva em:")
    print(f"    {FILA_PATH}")
    print(f"    {CMDS_PATH}")
    print(f"\n  Para executar: python auto_analisar_litispendencia.py")
    print()


def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--xlsx", default=str(DIR_FILES / "litispendencia.xlsx"),
                   help="Planilha de entrada (default: files/litispendencia.xlsx)")
    p.add_argument("--abas", default=",".join(ABAS_DEFAULT),
                   help="Abas a processar, separadas por vírgula")
    p.add_argument("--forcar", action="store_true",
                   help="Reprocessa mesmo grupos já em controle_grupos.json")
    args = p.parse_args()

    abas = [a.strip() for a in args.abas.split(",") if a.strip()]
    gerar_fila(Path(args.xlsx), abas, args.forcar)


if __name__ == "__main__":
    main()