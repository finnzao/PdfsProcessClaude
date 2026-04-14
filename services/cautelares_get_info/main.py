import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from common.fila_base import FilaBase
from common.consolidar_base import ConsolidarBase
from common.checkpoint import CheckpointManager
from common.sessao import SessaoManager
from common.utils import DIR_RESULT

SERVICE_DIR = Path(__file__).parent
RESULT_DIR = DIR_RESULT / "cautelares_get_info"


class FilaCustodiados(FilaBase):
    """Gera comandos para extração de dados pessoais dos réus."""
    SERVICE_NAME = "cautelares_get_info"
    BATCH_COM_PDF = 3
    BATCH_SEM_PDF = 10
    def _prompt_default(self): return "prompt_custodiado.md"
    def _prompt(self, cl): return "prompt_custodiado.md"

    def gerar_comando_com_pdf(self, n, procs, prompt):
        arqs = "\n".join(f"  - textos_extraidos/{p['txt_arquivo']}  ({p['numero']} | {p['classe']})" for p in procs)
        nums = " ".join(p["numero"] for p in procs)
        return f"""# === COMANDO {n:03d} === [CUSTODIADO] [{len(procs)} com PDF] ===
# Processos: {nums}
# Ao concluir: python run.py cautelares marcar {n} {nums}

Leia services/cautelares_get_info/prompts/{prompt}. Extraia dados de cada .md:

{arqs}

Salve em services/cautelares_get_info/resultados/custodiado_{n:03d}.json"""

    def gerar_comando_sem_pdf(self, n, procs, prompt):
        tab = "| Número | Classe | Assunto |\n|---|---|---|\n"
        tab += "\n".join(f"| {p['numero']} | {p['classe']} | {p['assunto']} |" for p in procs)
        nums = " ".join(p["numero"] for p in procs)
        return f"""# === COMANDO {n:03d} === [CUSTODIADO SEM PDF] ===
# Processos: {nums}
# Ao concluir: python run.py cautelares marcar {n} {nums}

Análise limitada:
{tab}

Salve em services/cautelares_get_info/resultados/custodiado_{n:03d}.json"""


class ConsolidarCustodiados(ConsolidarBase):
    """Junta JSONs e gera .xlsx para cadastro no ACLP."""
    SERVICE_NAME = "cautelares_get_info"

    def consolidar(self):
        print(f"{'='*60}\n  CONSOLIDACAO - Custodiados -> XLSX\n{'='*60}")
        jsons = sorted(self.resultados_dir.glob("custodiado_*.json"))
        if not jsons: print("  Sem resultados."); return
        todos = []
        for jp in jsons:
            try:
                d = json.loads(jp.read_text(encoding='utf-8'))
                if isinstance(d, list): todos.extend(d)
                elif "custodiados" in d: todos.extend(d["custodiados"])
                else: todos.append(d)
            except Exception as e: print(f"  AVISO {jp.name}: {e}")
        if not todos: print("  Vazio."); return
        print(f"  {len(todos)} custodiados")
        saida = self.result_dir / "custodiados_para_cadastro.xlsx"
        self._xlsx(todos, saida)
        print(f"  OK {saida}")

    def _xlsx(self, dados, path):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError: print("  pip install openpyxl"); return

        wb = Workbook(); ws = wb.active; ws.title = "Custodiados"
        hf = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        hfill = PatternFill('solid', fgColor='2F5496')
        ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
        am = PatternFill('solid', fgColor='FFF2CC')
        vm = PatternFill('solid', fgColor='F4CCCC')
        vd = PatternFill('solid', fgColor='D9EAD3')
        nf = Font(name='Arial', size=10)
        bd = Border(*[Side(style='thin', color='D9D9D9')]*4)
        PM = "PREENCHER MANUALMENTE"

        cols = [('A','Nome',25),('B','CPF',16),('C','RG',16),('D','Contato',18),('E','Processo',28),
            ('F','Vara',25),('G','Comarca',15),('H','Data Decisão',14),('I','Periodicidade',14),
            ('J','Data Compar.',16),('K','CEP',12),('L','Logradouro',30),('M','Nº',10),
            ('N','Complemento',15),('O','Bairro',20),('P','Cidade',15),('Q','UF',8),
            ('R','Comparecer',16),('S','Motivo',35),('T','Obs',40),('U','Doc',18)]
        for c, t, w in cols:
            cell = ws[f'{c}1']; cell.value = t; cell.font = hf; cell.fill = hfill; cell.alignment = ha; cell.border = bd
            ws.column_dimensions[c].width = w
        ws.freeze_panes = 'A2'

        for i, c in enumerate(dados, 2):
            cpf, rg = c.get('cpf',''), c.get('rg','')
            row = [c.get('nome',PM), cpf, rg, c.get('contato',PM), c.get('processo',''),
                c.get('vara','Vara Criminal de Rio Real'), c.get('comarca','Rio Real'),
                c.get('dataDecisao',PM), c.get('periodicidade',PM), c.get('dataComparecimentoInicial',PM),
                c.get('cep',PM), c.get('logradouro',PM), c.get('numero_endereco',''),
                c.get('complemento',''), c.get('bairro',PM), c.get('cidade','Rio Real'),
                c.get('estado','BA'), c.get('precisa_comparecer','VERIFICAR'),
                c.get('motivo_comparecimento',''), c.get('observacoes',''),
                'OK' if (cpf or rg) else 'SEM DOCUMENTO']
            for j, v in enumerate(row):
                cell = ws.cell(row=i, column=j+1, value=v); cell.font = nf; cell.border = bd
                if v == PM: cell.fill = am
                elif v == 'SEM DOCUMENTO': cell.fill = vm
                elif v == 'OK': cell.fill = vd
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)


def executar(comando, args=None):
    args = args or []
    if comando == "fila": FilaCustodiados(SERVICE_DIR).gerar()
    elif comando == "status": FilaCustodiados(SERVICE_DIR).status()
    elif comando == "analisar":
        SessaoManager(SERVICE_DIR/"checkpoint.json").inicio()
        print(f"  Cole comandos de: {SERVICE_DIR/'comandos_claude_code.txt'}")
    elif comando == "pausa": SessaoManager(SERVICE_DIR/"checkpoint.json").fim(SERVICE_DIR/"fila.json")
    elif comando == "marcar":
        if not args: print("  USO: marcar <NUM> [PROCESSOS...]"); return
        CheckpointManager(SERVICE_DIR/"checkpoint.json").marcar_concluido(int(args[0]), args[1:], f"resultados/custodiado_{int(args[0]):03d}.json")
    elif comando == "consolidar": ConsolidarCustodiados(SERVICE_DIR, RESULT_DIR).consolidar()
    elif comando == "reset":
        CheckpointManager(SERVICE_DIR/"checkpoint.json").reset()
        for f in [SERVICE_DIR/"fila.json", SERVICE_DIR/"comandos_claude_code.txt"]:
            if f.exists(): f.unlink()
        print("  Reset OK.")
    else: print(f"  Desconhecido: {comando}")
