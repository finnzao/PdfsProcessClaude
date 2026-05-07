"""
services/cautelares_get_info/scripts/consolidar_extracao.py — Consolida JSONs
extraídos pelo Claude Code em planilha xlsx para cadastro.

Diferença vs consolidar.py (regex pipeline): este lê o schema novo plano
gerado pelo Claude (array de objetos com campos diretos, sem aninhamento
em qualificacao/cautelar/metadados).

Schema esperado em resultados/extracao/extracao_NNN.json:
[
  {
    "numero_processo": "...",
    "nome": "...", "cpf": "...", "rg": "...",
    "telefone": "...",
    "cep": "...", "logradouro": "...", "numero_endereco": "...",
    "bairro": "...", "cidade": "...", "estado": "...",
    "status_cautelar": "ATIVA",
    "data_imposicao": "yyyy-MM-dd", "periodicidade_dias": 30,
    "peca_fonte": "...", "pagina_fonte": "...",
    "observacoes": "..."
  }
]
"""

import json
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

RAIZ = Path(__file__).resolve().parents[3]

EXTRACAO_DIR = RAIZ / "services" / "cautelares_get_info" / "resultados" / "extracao"
RESULT_DIR = RAIZ / "result" / "cautelares_get_info"


# ── Status de cadastro ──────────────────────────────────────────────

STATUS_PRONTO = "PRONTO"
STATUS_REVISAR = "REVISAR"
STATUS_BLOQUEADO = "BLOQUEADO"

CAUTELAR_CADASTRAR = {"ATIVA"}
CAUTELAR_REVISAR = {"SUSPEITA_ATIVA", "VERIFICAR"}
CAUTELAR_NAO_CADASTRAR = {
    "EXTINTA_REVOGACAO", "EXTINTA_CUMPRIMENTO",
    "EXTINTA_ABSOLVICAO", "EXTINTA_PUNIBILIDADE",
    "CONVERTIDA_PREVENTIVA", "NUNCA_IMPOSTA",
}


# ── Definição da planilha ───────────────────────────────────────────

# (nome_coluna, largura, grupo)
COLUNAS = [
    # Operacionais
    ("STATUS_CADASTRO",          12, "OP"),
    ("MOTIVO_REVISAO",           42, "OP"),

    # Identificação
    ("nome",                     32, "PESSOAL"),
    ("contato",                  16, "PESSOAL"),

    # Documentos
    ("cpf",                      16, "DOC"),
    ("rg",                       16, "DOC"),

    # Processo
    ("processo",                 26, "PROC"),
    ("vara",                     26, "PROC"),
    ("comarca",                  16, "PROC"),
    ("dataDecisao",              12, "PROC"),
    ("dataComparecimentoInicial",16, "PROC"),
    ("periodicidade",            10, "PROC"),

    # Endereço
    ("cep",                      11, "ENDR"),
    ("logradouro",               30, "ENDR"),
    ("numero",                    8, "ENDR"),
    ("complemento",              16, "ENDR"),
    ("bairro",                   18, "ENDR"),
    ("cidade",                   16, "ENDR"),
    ("estado",                    6, "ENDR"),

    # Observações (campo crítico do Claude)
    ("observacoes",              60, "OP"),

    # Diagnóstico cautelar (auxiliar)
    ("status_cautelar",          18, "AUX"),
    ("peca_fonte",               20, "AUX"),
    ("pagina_fonte",             14, "AUX"),
]

CORES_GRUPO = {
    "OP":      "595959",
    "PESSOAL": "1F3864",
    "DOC":     "2E75B6",
    "PROC":    "548235",
    "ENDR":    "C55A11",
    "AUX":     "7F6000",
}


# ── Normalizadores ──────────────────────────────────────────────────

def n_str(s, max_len: int = 200) -> str:
    if not s or not isinstance(s, str):
        return ""
    return re.sub(r"\s+", " ", s).strip()[:max_len]


def n_cpf(s) -> str:
    if not s or not isinstance(s, str):
        return ""
    digitos = re.sub(r"\D", "", s)
    if len(digitos) != 11:
        return ""
    return f"{digitos[:3]}.{digitos[3:6]}.{digitos[6:9]}-{digitos[9:]}"


def n_telefone(s) -> str:
    if not s or not isinstance(s, str):
        return "Pendente"
    permitidos = re.sub(r"[^\d() .\-]", "", s).strip()
    return permitidos if permitidos else "Pendente"


def n_cep(s) -> str:
    if not s or not isinstance(s, str):
        return ""
    digitos = re.sub(r"\D", "", s)
    if len(digitos) != 8:
        return ""
    return f"{digitos[:5]}-{digitos[5:]}"


UFS_VALIDAS = {
    "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS",
    "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC",
    "SP", "SE", "TO",
}

ESTADOS_POR_EXTENSO = {
    "ACRE": "AC", "ALAGOAS": "AL", "AMAPA": "AP", "AMAZONAS": "AM",
    "BAHIA": "BA", "CEARA": "CE", "DISTRITO FEDERAL": "DF",
    "ESPIRITO SANTO": "ES", "GOIAS": "GO", "MARANHAO": "MA",
    "MATO GROSSO": "MT", "MATO GROSSO DO SUL": "MS", "MINAS GERAIS": "MG",
    "PARA": "PA", "PARAIBA": "PB", "PARANA": "PR", "PERNAMBUCO": "PE",
    "PIAUI": "PI", "RIO DE JANEIRO": "RJ", "RIO GRANDE DO NORTE": "RN",
    "RIO GRANDE DO SUL": "RS", "RONDONIA": "RO", "RORAIMA": "RR",
    "SANTA CATARINA": "SC", "SAO PAULO": "SP", "SERGIPE": "SE",
    "TOCANTINS": "TO",
}


def n_estado(s) -> str:
    """
    Normaliza estado para sigla UF de 2 letras maiúsculas.
    Aceita: 'BA', 'ba', 'Bahia', 'BAHIA', 'B.A.', 'Rio de Janeiro', etc.
    Retorna '' se não conseguir mapear para uma UF válida.
    """
    if not s or not isinstance(s, str):
        return ""
    # Remove acentos via mapeamento simples
    acentos = str.maketrans("ÁÀÂÃÄÉÈÊËÍÌÎÏÓÒÔÕÖÚÙÛÜÇ", "AAAAAEEEEIIIIOOOOOUUUUC")
    s_norm = s.upper().translate(acentos).strip()

    # Tenta como nome por extenso
    s_clean = re.sub(r"[^A-Z\s]", "", s_norm).strip()
    s_clean = re.sub(r"\s+", " ", s_clean)
    if s_clean in ESTADOS_POR_EXTENSO:
        return ESTADOS_POR_EXTENSO[s_clean]

    # Tenta como sigla — pega só letras e checa as 2 primeiras
    s_letras = re.sub(r"[^A-Z]", "", s_norm)
    if len(s_letras) >= 2:
        sigla = s_letras[:2]
        if sigla in UFS_VALIDAS:
            return sigla

    return ""


def n_data(s) -> str:
    """Para yyyy-MM-dd. Aceita ISO direto ou converte de dd/mm/yyyy."""
    if not s or not isinstance(s, str):
        return ""
    s = s.strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s
    m = re.match(r"^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})$", s)
    if m:
        d, mo, a = m.groups()
        if len(a) == 2:
            a = ("20" + a) if int(a) < 50 else ("19" + a)
        try:
            return f"{int(a):04d}-{int(mo):02d}-{int(d):02d}"
        except ValueError:
            pass
    return ""


def n_processo(s) -> str:
    if not s or not isinstance(s, str):
        return ""
    return re.sub(r"[^\d.\-]", "", s)


def n_periodicidade(p) -> int | None:
    if isinstance(p, int) and 1 <= p <= 365:
        return p
    if isinstance(p, str):
        digitos = re.sub(r"\D", "", p)
        if digitos and 1 <= int(digitos) <= 365:
            return int(digitos)
    return None


def calcular_comparecimento_inicial(data_iso: str, dias: int | None) -> str:
    if not data_iso or not dias:
        return ""
    try:
        dt = datetime.strptime(data_iso, "%Y-%m-%d").date()
        return (dt + timedelta(days=dias)).strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return ""


# ── Construção e validação da linha ─────────────────────────────────

def construir_linha(reg: dict) -> dict:
    data_dec = n_data(reg.get("data_imposicao"))
    per_dias = n_periodicidade(reg.get("periodicidade_dias"))
    data_comp = calcular_comparecimento_inicial(data_dec, per_dias)

    return {
        "nome": n_str(reg.get("nome"), 150),
        "contato": n_telefone(reg.get("telefone")),
        "cpf": n_cpf(reg.get("cpf")),
        "rg": n_str(reg.get("rg"), 20),
        "processo": n_processo(reg.get("numero_processo")),
        "vara": n_str(reg.get("vara") or "Vara Criminal de Rio Real", 100),
        "comarca": "Rio Real",
        "dataDecisao": data_dec,
        "dataComparecimentoInicial": data_comp,
        "periodicidade": per_dias if per_dias else "",
        "cep": n_cep(reg.get("cep")),
        "logradouro": n_str(reg.get("logradouro"), 200),
        "numero": n_str(reg.get("numero_endereco"), 20),
        "complemento": n_str(reg.get("complemento"), 100),
        "bairro": n_str(reg.get("bairro"), 100),
        "cidade": n_str(reg.get("cidade") or "Rio Real", 100),
        "estado": n_estado(reg.get("estado") or "BA"),
        "observacoes": n_str(reg.get("observacoes"), 1000),
        "status_cautelar": n_str(reg.get("status_cautelar"), 25),
        "peca_fonte": n_str(reg.get("peca_fonte"), 30),
        "pagina_fonte": n_str(reg.get("pagina_fonte"), 20),
    }


def validar(linha: dict) -> tuple[str, list[str]]:
    """
    Validação simplificada: só checa presença dos campos obrigatórios e
    se o estado está em sigla de 2 letras. Tamanhos de string não bloqueiam.
    """
    motivos = []
    status_cau = linha["status_cautelar"]

    # Bloqueios por status de cautelar
    if status_cau in CAUTELAR_NAO_CADASTRAR:
        motivos.append(f"cautelar {status_cau} — não cadastrar")
        return STATUS_BLOQUEADO, motivos

    # ── Campos obrigatórios (só presença) ──
    if not linha["nome"]:
        motivos.append("nome ausente")
    if not linha["cpf"] and not linha["rg"]:
        motivos.append("sem CPF nem RG")
    if not linha["processo"]:
        motivos.append("processo ausente")
    if not linha["dataDecisao"]:
        motivos.append("dataDecisao ausente")
    if not linha["periodicidade"]:
        motivos.append("periodicidade ausente")
    if not linha["cep"]:
        motivos.append("cep ausente")
    if not linha["logradouro"]:
        motivos.append("logradouro ausente")
    if not linha["bairro"]:
        motivos.append("bairro ausente")
    if not linha["cidade"]:
        motivos.append("cidade ausente")

    # ── Estado: precisa ser sigla de 2 letras maiúsculas ──
    if not linha["estado"] or not re.match(r"^[A-Z]{2}$", linha["estado"]):
        motivos.append("estado precisa ser sigla com 2 letras (ex: BA, SP)")

    # Qualquer ausência de obrigatório bloqueia
    if motivos:
        return STATUS_BLOQUEADO, motivos

    # ── Avisos (passa, mas merece olhar) ──
    if status_cau in CAUTELAR_REVISAR:
        motivos.append(f"cautelar {status_cau} — verificar")
    if linha["contato"] == "Pendente":
        motivos.append("telefone não localizado")
    if not linha["cpf"]:
        motivos.append("sem CPF (apenas RG)")
    # dataComparecimentoInicial é OPCIONAL — não vira aviso
    if linha["observacoes"] and len(linha["observacoes"]) > 30:
        motivos.append("Claude registrou gaps — ler observações")

    if motivos:
        return STATUS_REVISAR, motivos
    return STATUS_PRONTO, []


# ── Pipeline ────────────────────────────────────────────────────────

def carregar_extracoes(extracao_dir: Path) -> list[dict]:
    arquivos = sorted(extracao_dir.glob("extracao_*.json"))
    if not arquivos:
        return []

    todos = []
    vistos = set()
    for arq in arquivos:
        try:
            dados = json.loads(arq.read_text(encoding="utf-8"))
            if not isinstance(dados, list):
                if isinstance(dados, dict):
                    dados = [dados]
                else:
                    print(f"  ⚠️  {arq.name}: formato inesperado, ignorado")
                    continue
            for d in dados:
                if not isinstance(d, dict):
                    continue
                # Deduplica por (processo, nome) — múltiplos réus são linhas distintas
                chave = (d.get("numero_processo", ""), d.get("nome", ""))
                if chave in vistos:
                    continue
                vistos.add(chave)
                todos.append(d)
        except Exception as e:
            print(f"  ⚠️  {arq.name}: {e}")

    return todos


def gerar_xlsx(linhas: list[dict], saida: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print("  ✗ openpyxl não instalado: pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Cadastro"

    cab_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cab_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Arial", size=9)
    cell_align = Alignment(vertical="top", wrap_text=True)
    borda = Border(*[Side(style="thin", color="D9D9D9")] * 4)

    cores_status = {
        STATUS_PRONTO:    PatternFill("solid", fgColor="D9EAD3"),
        STATUS_REVISAR:   PatternFill("solid", fgColor="FFF2CC"),
        STATUS_BLOQUEADO: PatternFill("solid", fgColor="F4CCCC"),
    }
    fontes_status = {
        STATUS_PRONTO:    Font(name="Arial", size=9, bold=True, color="274E13"),
        STATUS_REVISAR:   Font(name="Arial", size=9, bold=True, color="7F6000"),
        STATUS_BLOQUEADO: Font(name="Arial", size=9, bold=True, color="990000"),
    }

    # Cabeçalho
    for ci, (nome, larg, grupo) in enumerate(COLUNAS, 1):
        c = ws.cell(row=1, column=ci, value=nome)
        c.font = cab_font
        c.fill = PatternFill("solid", fgColor=CORES_GRUPO[grupo])
        c.alignment = cab_align
        c.border = borda
        ws.column_dimensions[get_column_letter(ci)].width = larg
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "C2"

    if linhas:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS))}{len(linhas) + 1}"

    nomes_cols = [c[0] for c in COLUNAS]
    col_status = nomes_cols.index("STATUS_CADASTRO") + 1
    col_estado = nomes_cols.index("estado") + 1
    col_per = nomes_cols.index("periodicidade") + 1

    dv_status = DataValidation(
        type="list",
        formula1=f'"{STATUS_PRONTO},{STATUS_REVISAR},{STATUS_BLOQUEADO}"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_status)

    UFS = "AC,AL,AP,AM,BA,CE,DF,ES,GO,MA,MT,MS,MG,PA,PB,PR,PE,PI,RJ,RN,RS,RO,RR,SC,SP,SE,TO"
    dv_uf = DataValidation(type="list", formula1=f'"{UFS}"', allow_blank=False)
    ws.add_data_validation(dv_uf)

    for ri, linha in enumerate(linhas, 2):
        for ci, nome in enumerate(nomes_cols, 1):
            v = linha.get(nome, "")
            if v == "" or v is None:
                v = None
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = cell_font
            c.alignment = cell_align
            c.border = borda

        per = linha.get("periodicidade")
        if isinstance(per, int):
            ws.cell(row=ri, column=col_per).number_format = "0"

        st = linha.get("STATUS_CADASTRO", "")
        if st in cores_status:
            ws.cell(row=ri, column=col_status).fill = cores_status[st]
            ws.cell(row=ri, column=col_status).font = fontes_status[st]

        dv_status.add(ws.cell(row=ri, column=col_status).coordinate)
        dv_uf.add(ws.cell(row=ri, column=col_estado).coordinate)

    saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(saida)


def consolidar(extracao_dir: Path | None = None, saida: Path | None = None) -> None:
    extracao_dir = Path(extracao_dir) if extracao_dir else EXTRACAO_DIR
    saida = Path(saida) if saida else RESULT_DIR / "custodiados_cadastro.xlsx"

    print(f"  Lendo: {extracao_dir}")
    registros = carregar_extracoes(extracao_dir)
    if not registros:
        print(f"  ✗ Nenhuma extração encontrada.")
        print(f"     Rode primeiro: python auto_extrair_cautelares.py")
        return

    print(f"  {len(registros)} registros encontrados\n")

    linhas = []
    for reg in registros:
        linha = construir_linha(reg)
        status, motivos = validar(linha)
        linha["STATUS_CADASTRO"] = status
        linha["MOTIVO_REVISAO"] = "; ".join(motivos)
        linhas.append(linha)

    gerar_xlsx(linhas, saida)

    # Resumo
    n_pronto = sum(1 for l in linhas if l["STATUS_CADASTRO"] == STATUS_PRONTO)
    n_revisar = sum(1 for l in linhas if l["STATUS_CADASTRO"] == STATUS_REVISAR)
    n_bloq = sum(1 for l in linhas if l["STATUS_CADASTRO"] == STATUS_BLOQUEADO)

    print(f"  ── Resumo ──")
    print(f"  Total:      {len(linhas)}")
    print(f"  ✓ PRONTO:   {n_pronto:>3}  (importador consome direto)")
    print(f"  ⚠ REVISAR:  {n_revisar:>3}  (humano analisa antes)")
    print(f"  ✗ BLOQUEADO:{n_bloq:>3}  (descartado)")
    print(f"\n  ✓ Planilha: {saida}\n")


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--src", help="Pasta com JSONs de extração")
    p.add_argument("--saida", help="Caminho da planilha xlsx")
    args = p.parse_args()
    consolidar(args.src, args.saida)
