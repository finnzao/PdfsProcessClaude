"""
test_consolidador.py — Validação rápida do consolidador alinhado à DTO.
Cria 4 JSONs sintéticos (PRONTO, REVISAR, BLOQUEADO×2) e roda o pipeline.

Roda a partir da raiz do projeto:
    python -m services.cautelares_get_info.scripts.tests.test_consolidador
ou diretamente:
    python services/cautelares_get_info/scripts/tests/test_consolidador.py
"""

import json
import sys
import tempfile
from pathlib import Path

# Adiciona a raiz do projeto ao sys.path para que `from services...` funcione
# quando o teste é executado diretamente (não via -m)
RAIZ_PROJETO = Path(__file__).resolve().parents[4]
if str(RAIZ_PROJETO) not in sys.path:
    sys.path.insert(0, str(RAIZ_PROJETO))

from services.cautelares_get_info.scripts.consolidar import consolidar


# Caso 1: completo, deveria virar PRONTO
JSON_PRONTO = {
    "numero_processo": "8001234-56.2024.8.05.0216",
    "metadados_processo": {
        "classe": "Ação Penal",
        "assunto": "Roubo Majorado",
        "orgao_julgador": "Vara Criminal de Rio Real",
    },
    "qualificacao": {
        "nome": "JOÃO DA SILVA SANTOS",
        "cpf": "123.456.789-00",
        "rg": "12.345.678 SSP/BA",
        "data_nascimento": "15/05/1990",
        "telefone": "(75) 99999-1234",
        "cep": "48340-000",
        "logradouro": "Rua das Flores",
        "numero_endereco": "123",
        "bairro": "Centro",
        "cidade": "Rio Real",
        "estado": "BA",
        "nacionalidade": "Brasileira",
        "estado_civil": "solteiro",
    },
    "cautelar": {
        "status": "ATIVA",
        "imposta": True,
        "peca_fonte": "AUDIENCIA_CUSTODIA",
        "pagina_fonte": "p.10-15",
        "doc_id_fonte": "Num. 440867200",
        "data_imposicao": "16/03/2024",
        "periodicidade": "mensal",
        "condicoes": ["Comparecimento periódico", "Proibição de contato"],
        "confianca": "alta",
        "motivo_status": "Cautelar imposta em 16/03/2024",
        "sinalizadores": [],
    },
    "needs_llm": False,
}

# Caso 2: telefone ausente + status SUSPEITA_ATIVA → REVISAR
JSON_REVISAR = {
    "numero_processo": "8005678-12.2023.8.05.0216",
    "metadados_processo": {
        "classe": "Ação Penal",
        "assunto": "Furto",
        "orgao_julgador": "Vara Criminal de Rio Real",
    },
    "qualificacao": {
        "nome": "MARIA APARECIDA OLIVEIRA",
        "cpf": "987.654.321-00",
        "rg": "98.765.432 SSP/BA",
        "data_nascimento": "20/08/1985",
        "telefone": "",  # vazio
        "cep": "48340-000",
        "logradouro": "Rua dos Jasmins",
        "numero_endereco": "45",
        "bairro": "Centro",
        "cidade": "Rio Real",
        "estado": "BA",
    },
    "cautelar": {
        "status": "SUSPEITA_ATIVA",
        "imposta": True,
        "peca_fonte": "SURSIS_PROCESSUAL",
        "pagina_fonte": "p.30",
        "data_imposicao": "10/01/2022",
        "periodicidade": "mensal",
        "periodo_prova": "2 anos",
        "condicoes": ["Comparecimento periódico"],
        "confianca": "baixa",
        "motivo_status": "Sursis homologado, sem certidão de cumprimento",
        "sinalizadores": ["Verificar período de prova"],
    },
    "needs_llm": True,
    "campos_para_revisao_llm": [
        {"campo": "status_cautelar", "motivo": "verificar cumprimento", "prioridade": "critica"}
    ],
}

# Caso 3: sem documentos → BLOQUEADO
JSON_BLOQUEADO = {
    "numero_processo": "8009999-99.2024.8.05.0216",
    "metadados_processo": {
        "classe": "Ação Penal",
        "orgao_julgador": "Vara Criminal de Rio Real",
    },
    "qualificacao": {
        "nome": "PEDRO HENRIQUE",
        "cpf": "",
        "rg": "",
        "data_nascimento": "",
        "cep": "48340-000",
        "logradouro": "Rua",
        "bairro": "Centro",
        "cidade": "Rio Real",
        "estado": "BA",
    },
    "cautelar": {
        "status": "ATIVA",
        "imposta": True,
        "peca_fonte": "AUDIENCIA_CUSTODIA",
        "data_imposicao": "01/02/2024",
        "periodicidade": "mensal",
        "confianca": "alta",
        "motivo_status": "",
        "condicoes": [],
        "sinalizadores": [],
    },
    "needs_llm": True,
}

# Caso 4: cautelar EXTINTA → BLOQUEADO mesmo com dados completos
JSON_EXTINTO = {
    "numero_processo": "8002000-00.2020.8.05.0216",
    "metadados_processo": {
        "classe": "Ação Penal",
        "orgao_julgador": "Vara Criminal de Rio Real",
    },
    "qualificacao": {
        "nome": "CARLOS ALBERTO",
        "cpf": "111.222.333-44",
        "rg": "11.222.333 SSP/BA",
        "data_nascimento": "01/01/1980",
        "telefone": "(75) 88888-5555",
        "cep": "48340-000",
        "logradouro": "Av Principal",
        "numero_endereco": "200",
        "bairro": "Centro",
        "cidade": "Rio Real",
        "estado": "BA",
    },
    "cautelar": {
        "status": "EXTINTA_CUMPRIMENTO",
        "imposta": True,
        "peca_fonte": "SURSIS_PROCESSUAL",
        "data_imposicao": "01/01/2020",
        "periodicidade": "mensal",
        "confianca": "alta",
        "motivo_status": "Sursis cumprido",
        "condicoes": [],
        "sinalizadores": [],
    },
    "needs_llm": False,
}


def main():
    with tempfile.TemporaryDirectory() as tmp:
        json_dir = Path(tmp) / "pre_extraido"
        json_dir.mkdir()

        casos = {
            "caso_pronto.json": JSON_PRONTO,
            "caso_revisar.json": JSON_REVISAR,
            "caso_bloqueado.json": JSON_BLOQUEADO,
            "caso_extinto.json": JSON_EXTINTO,
        }
        for nome, payload in casos.items():
            (json_dir / nome).write_text(
                json.dumps(payload, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

        saida = Path(tmp) / "cadastro.xlsx"
        consolidar(json_dir, None, saida)

        # Verifica
        from openpyxl import load_workbook
        wb = load_workbook(saida)
        ws = wb["Cadastro"]
        cabecalho = [c.value for c in ws[1]]

        print("\n  ── Cabeçalho da planilha ──")
        for nome in cabecalho:
            print(f"    {nome}")

        print("\n  ── Linhas geradas ──")
        idx_status = cabecalho.index("STATUS_CADASTRO")
        idx_nome = cabecalho.index("nome")
        idx_motivo = cabecalho.index("MOTIVO_REVISAO")
        idx_per = cabecalho.index("periodicidade")
        idx_cpf = cabecalho.index("cpf")
        idx_data = cabecalho.index("dataDecisao")
        idx_inicial = cabecalho.index("dataComparecimentoInicial")

        for row in ws.iter_rows(min_row=2, values_only=True):
            print(f"    {str(row[idx_status]):<10} | {row[idx_nome][:25]:<25} | "
                  f"CPF:{str(row[idx_cpf]):<16} | per:{str(row[idx_per]):<3} | "
                  f"dec:{str(row[idx_data]):<10} | ini:{str(row[idx_inicial]):<10}")
            motivo = row[idx_motivo]
            if motivo:
                print(f"    {'':10}   motivo: {motivo}")

        # Asserts
        linhas_status = [r[idx_status] for r in ws.iter_rows(min_row=2, values_only=True)]
        assert "PRONTO" in linhas_status, "deveria ter ao menos uma linha PRONTO"
        assert "REVISAR" in linhas_status, "deveria ter ao menos uma linha REVISAR"
        assert "BLOQUEADO" in linhas_status, "deveria ter ao menos uma linha BLOQUEADO"
        assert linhas_status.count("BLOQUEADO") == 2, "esperava 2 BLOQUEADO (sem doc + extinto)"

        print("\n  ✓ Todos os asserts passaram")


if __name__ == "__main__":
    main()
